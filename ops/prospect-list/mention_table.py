"""Count who the assistants named, per assistant, and sort the outreach list by it.

Step 5 of the method in `ops/outreach.md` section 3 — the last cut, and the one
that decides which draft each practice gets.

    python3 mention_table.py run.csv locations.csv out.xlsx [outreach.txt]

  run.csv        a trade run from `ops/trade-run/trade_run.py`
  locations.csv  the CQC list from `cqc_filter.py` — the candidate names
  out.xlsx       written outside this repository, per the .gitignore here
  outreach.txt   optional: one practice name per line, the triaged list. Given
                 one, the workbook gets a third sheet holding only those, which
                 is the sheet you actually send from. Without it you get the
                 whole market, which is the sheet you study.

**Score the whole market, send to the triaged few.** The two are different jobs
and the file does both. Counting only the approachable practices would hide the
competitors that make the email worth reading, since the businesses ranked ahead
of a prospect are usually the ones we cannot or should not sell to.

**The candidate list is the point.** We are not discovering which practices
exist, we are checking which of a known set got named, which is why this counts
rather than guesses. Names are matched against `answer_text` directly, because
`trade_run.py` writes the `competitors` column empty by design.

**Two numbers, and they are not interchangeable.** Mentions — how often a
practice is named in an answer — can be counted for all three assistants.
Sources — which pages the answer was built from — can only be counted for
ChatGPT and Perplexity, because every one of Gemini's cited URLs is an opaque
`vertexaisearch` redirect (`ops/competitor-analysis.md` Finding E). Any
conclusion drawn from the source table has to say so.

**It prints what it could not decide.** A practice whose name never matched, and
a name that only matched loosely, are both listed at the end rather than silently
scored zero. A zero that is really a matching failure would put a practice in
Draft B and make the email wrong.
"""

import csv
import re
import sys
from collections import Counter, defaultdict

# Words that carry no distinguishing information in a dental practice's name.
# Stripped to build the short alias; never used as an alias on their own.
NOISE = {
    'the', 'dental', 'dentist', 'dentists', 'dentistry', 'practice', 'practices',
    'surgery', 'surgeries', 'clinic', 'clinics', 'centre', 'center', 'studio',
    'care', 'group', 'limited', 'ltd', 'llp', 'and', 'of', 'ltd.', 'co',
    'orthodontics', 'orthodontic', 'implants', 'aesthetic', 'aesthetics',
    'wirral', 'merseyside',
}

ASSISTANTS = ['chatgpt', 'gemini', 'perplexity']
LABEL = {'chatgpt': 'ChatGPT', 'gemini': 'Gemini', 'perplexity': 'Perplexity'}


def label(assistant):
    return LABEL.get(assistant, assistant.title())


def norm(text):
    """Lowercase, collapse punctuation to single spaces, keep digits."""
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9]+', ' ', (text or '').lower())).strip()


def aliases(name):
    """Full normalised name, plus the distinctive core of it.

    "The Sandstone Dental Practice" gives "the sandstone dental practice" and
    "sandstone". "No.22 Implants and Aesthetic Dentistry" gives the full string
    and "22" — which is too short to be safe, so short cores are dropped and the
    practice falls to the review list instead of being scored on a false match.
    """
    full = norm(name)
    core = ' '.join(w for w in full.split() if w not in NOISE)
    out = {full}
    if core and len(core) >= 5 and core != full:
        out.add(core)
    return out, core


def domain(url):
    m = re.match(r'https?://([^/]+)', (url or '').strip(), re.I)
    if not m:
        return ''
    return re.sub(r'^www\.', '', m.group(1).lower())


def load_run(path):
    rows = list(csv.DictReader(open(path, encoding='utf-8-sig')))
    for r in rows:
        r['_assistant'] = norm(r.get('assistant'))
    bad = [r for r in rows if (r.get('errors') or '').strip()]
    if bad:
        print(f'note: {len(bad)} rows carry an error and are excluded from the counts',
              file=sys.stderr)
    return [r for r in rows if not (r.get('errors') or '').strip()]


def main():
    run_path, loc_path, out_path = sys.argv[1:4]
    outreach_path = sys.argv[4] if len(sys.argv) > 4 else None
    run = load_run(run_path)
    locs = list(csv.DictReader(open(loc_path, encoding='utf-8-sig')))

    outreach = set()
    if outreach_path:
        with open(outreach_path, encoding='utf-8-sig') as f:
            outreach = {norm(l) for l in f if l.strip() and not l.startswith('#')}

    seen_assistants = sorted({r['_assistant'] for r in run})
    known = [a for a in ASSISTANTS if a in seen_assistants]
    extra = [a for a in seen_assistants if a not in ASSISTANTS]
    if extra:
        print(f'note: unexpected assistant names in the run: {extra}', file=sys.stderr)
        known += extra
    rows_per = Counter(r['_assistant'] for r in run)

    # One normalised haystack per row, built once.
    for r in run:
        r['_hay'] = norm(r.get('answer_text'))

    # A site that changed hands is registered twice, once under each provider.
    # Scoring it twice would double it in the sort, so it is collapsed to one row
    # and the providers are shown together.
    merged, order_seen = {}, []
    for loc in locs:
        name = (loc.get('Name') or '').strip()
        if not name:
            continue
        # Keyed on the postcode too: two sites can share a name and be different
        # businesses, and merging those would delete a prospect.
        key = (name, (loc.get('Postcode') or '').strip())
        if key not in merged:
            merged[key] = loc
            order_seen.append(key)
            merged[key]['_providers'] = [loc.get('Provider name', '')]
        else:
            merged[key]['_providers'].append(loc.get('Provider name', ''))
    dupes = [k for k in order_seen if len(merged[k]['_providers']) > 1]

    # A name at two postcodes is two businesses that the answers cannot tell
    # apart, so both carry the same count and a person has to split them.
    by_name = Counter(k[0] for k in order_seen)
    collisions = sorted({n for n, c in by_name.items() if c > 1})

    table, unmatched, weak = [], [], []
    for key in order_seen:
        name = key[0]
        loc = merged[key]
        keys, core = aliases(name)
        counts = {a: 0 for a in known}
        for r in run:
            if any(k in r['_hay'] for k in keys):
                counts[r['_assistant']] = counts.get(r['_assistant'], 0) + 1
        total = sum(counts.values())
        naming = sum(1 for a in known if counts[a] > 0)
        missing = [a for a in known if counts[a] == 0]

        if total == 0:
            unmatched.append(name)
        elif core and len(core) < 5:
            weak.append(name)

        row = {
            'Practice': name,
            'Postcode': loc.get('Postcode', ''),
            'Provider': ' / '.join(dict.fromkeys(loc['_providers'])),
        }
        for a in known:
            n = rows_per[a]
            row[label(a)] = counts[a]
            row[f'{label(a)} %'] = round(100 * counts[a] / n) if n else 0
        row['Total'] = total
        row['% of all answers'] = round(100 * total / len(run)) if run else 0
        row['Assistants naming'] = naming
        row['Missing from'] = ', '.join(label(m) for m in missing) or '—'
        row['_outreach'] = norm(name) in outreach

        # The visibility ladder, ops/outreach.md section 4.
        if naming == 0:
            row['Draft'], row['Why'] = 'B', 'Named by none of the three.'
        elif naming == len(known):
            row['Draft'] = 'Do not contact'
            row['Why'] = ('Named by all three. The ladder says there is nothing to sell '
                          'them. If the counts are low, look before you drop them.')
        else:
            row['Draft'] = 'A'
            row['Why'] = f'Named by {naming} of {len(known)}, missing from ' \
                         f'{row["Missing from"]}.'
        table.append(row)

    # Sort: the practices with a real gap first, then by how big the gap is.
    order = {'A': 0, 'B': 1, 'Do not contact': 2}
    table.sort(key=lambda r: (order[r['Draft']], -r['Total'], r['Practice']))

    # Sources: ChatGPT and Perplexity only. Gemini's URLs cannot be resolved.
    src = defaultdict(Counter)
    opaque = 0
    for r in run:
        a = r['_assistant']
        urls = [u for u in re.split(r'[\s,;|]+', r.get('sources_cited') or '') if u]
        for u in urls:
            d = domain(u)
            if not d:
                continue
            if 'vertexaisearch' in d:
                opaque += 1
                continue
            src[a][d] += 1

    send = [r for r in table if r['_outreach']] if outreach else []
    if outreach:
        matched = {norm(r['Practice']) for r in send}
        missed = sorted(outreach - matched)
        if missed:
            print(f'\nOn the outreach list but not in the CQC file — spelling, or a '
                  f'practice that has deregistered ({len(missed)}):', file=sys.stderr)
            for n in missed:
                print(f'  {n}', file=sys.stderr)

    write(out_path, table, known, rows_per, len(run), src, opaque, send)

    print(f'\n{len(table)} practices scored across {len(run)} answers -> {out_path}',
          file=sys.stderr)
    counts = Counter(r['Draft'] for r in table)
    print(f"  Draft A {counts['A']}, Draft B {counts['B']}, "
          f"do not contact {counts['Do not contact']}", file=sys.stderr)
    if unmatched:
        print(f'\nNever matched, so scored zero — check these by hand before trusting '
              f'a Draft B ({len(unmatched)}):', file=sys.stderr)
        for n in unmatched:
            print(f'  {n}', file=sys.stderr)
    if weak:
        print(f'\nName too short to match safely, counted on the full name only '
              f'({len(weak)}):', file=sys.stderr)
        for n in weak:
            print(f'  {n}', file=sys.stderr)
    if dupes:
        print(f'\nRegistered twice and merged into one row — confirm which provider '
              f'is current ({len(dupes)}):', file=sys.stderr)
        for k in dupes:
            print(f'  {k[0]} {k[1]}: '
                  f'{" / ".join(dict.fromkeys(merged[k]["_providers"]))}', file=sys.stderr)
    if collisions:
        print(f'\nSame name at more than one postcode — the answers cannot tell these '
              f'apart, so both rows carry the same count ({len(collisions)}):',
              file=sys.stderr)
        for n in collisions:
            print(f'  {n}', file=sys.stderr)


def write(path, table, known, rows_per, total_rows, src, opaque, send):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY, INK, BRASS, WARM = '170969', '16161D', '8A6A28', 'FFFEFA'
    RULE = Side(style='thin', color='D8D4CA')

    def sheet(ws, title, intro, headers, rows, widths, wrap_from):
        ws.sheet_view.showGridLines = False
        ws['A1'] = title
        ws['A1'].font = Font(name='Palatino Linotype', size=16, bold=True, color=NAVY)
        ws['A2'] = intro
        ws['A2'].font = Font(name='Segoe UI', size=9, italic=True, color='5B5B66')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[2].height = 30
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = Font(name='Segoe UI', size=9, bold=True, color=WARM)
            cell.fill = PatternFill('solid', fgColor=NAVY)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[4].height = 26
        for i, row in enumerate(rows, 5):
            for c, h in enumerate(headers, 1):
                v = row.get(h, '') if isinstance(row, dict) else row[c - 1]
                cell = ws.cell(row=i, column=c, value=v)
                cell.font = Font(name='Segoe UI', size=9, color=INK)
                cell.alignment = Alignment(vertical='top', wrap_text=(c >= wrap_from))
                cell.border = Border(bottom=RULE)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = ws.cell(row=5, column=1)

    head = ['Practice', 'Postcode', 'Provider']
    for a in known:
        head += [label(a), f'{label(a)} %']
    head += ['Total', '% of all answers', 'Assistants naming', 'Missing from',
             'Draft', 'Why']
    widths = [34, 10, 32] + [9, 9] * len(known) + [8, 10, 11, 20, 14, 46]

    wb = Workbook()
    per = ', '.join(f'{label(a)} {rows_per[a]}' for a in known)

    if send:
        ws = wb.active
        ws.title = 'Send from this'
        sheet(ws, 'The outreach list, sorted by who named them',
              f'Only the practices on the triaged outreach list. Draft A at the top and '
              f'the biggest gap first, because a practice one assistant names well and '
              f'another misses entirely is the sharpest email we can write. Counts come '
              f'from {total_rows} answers ({per}).',
              head, send, widths, len(head))
        wb.create_sheet('Whole market')

    sheet(wb['Whole market'] if send else wb.active,
          'Wirral dentists, sorted by who named them',
          f'{total_rows} answers in the run ({per}). Draft A first and the biggest gaps '
          f'at the top, then Draft B, then the practices already named everywhere. '
          f'A percentage is of that assistant\'s own answers, so the columns are '
          f'comparable to each other.',
          head, table, widths, len(head))
    if not send:
        wb.active.title = 'Whole market'

    ws = wb.create_sheet('Sources')
    src_rows = []
    for a in sorted(src):
        for d, n in src[a].most_common(40):
            src_rows.append([label(a), d, n])
    sheet(ws, 'Where the answers came from',
          'ChatGPT and Perplexity only. Every one of Gemini\'s cited URLs is an opaque '
          f'vertexaisearch redirect — {opaque} of them in this run — so no source '
          'analysis is possible for it, and any conclusion here rests on two assistants '
          'rather than three.',
          ['Assistant', 'Domain', 'Rows citing it'], src_rows, [16, 44, 14], 4)

    wb.save(path)


if __name__ == '__main__':
    main()
