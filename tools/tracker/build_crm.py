#!/usr/bin/env python3
"""Build wardith-crm.xlsx, the daily-operating-system workbook, from
tracker.json plus (if it already exists) the previously-built workbook's
own Activities log and manual fields - which are never regenerated, only
carried forward. See tools/tracker/README.md for the full picture and
tools/tracker/cadence.py for the cadence/stage rules every formula in this
file reproduces.

Usage:
  python3 build_crm.py [--tracker-dir DIR] [--output PATH]

Default tracker dir: ~/wardith-runs/tracker (same default as
import_tracker.py). Default output: <tracker-dir>/wardith-crm.xlsx.

Read-only against tracker.json/tracker.csv and the existing workbook (if
present); the only write is the new workbook, via a temp-file-then-replace
so a crash mid-build never leaves a half-written .xlsx in place of a good
one.
"""
import argparse
import json
import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import xlsxwriter
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import cadence as cad

# ---------------------------------------------------------------------------
# Brand palette - playbook/site-code-locations.md / site/src/styles/global.css.
# Navy #170969, warm white #fffefa: the same colours the published site uses,
# not a generic spreadsheet blue.
# ---------------------------------------------------------------------------
NAVY = "#170969"
NAVY_DEEP = "#0C0438"
PAPER = "#FFFEFA"
PAPER_2 = "#F7F5EF"
INK = "#16161D"
RULE = "#E2DFD6"
BRASS = "#8A6A28"

OVERDUE_BG = "#F8D7DA"
DUE_TODAY_BG = "#FFF3CD"
REPLIED_BG = "#CFE2FF"
WON_BG = "#C6EFCE"
LOST_BG = "#E2E2E2"
REVIEW_BG = "#FCE4D6"
BLOCKED_BG = "#D9D9D9"
READY_BG = "#D4EDDA"

PLAN_PRICES = [("Maintain", 150), ("Grow", 400), ("Lead", 700)]
AUDIT_PRICE = 250
FOUNDATION_PRICE = 800

PROSPECTS_COLUMNS = [
    ("prospect_id", "Prospect ID", 34),
    ("business", "Business", 28),
    ("contact_person", "Contact Person", 20),
    ("contact_email", "Contact Email", 24),
    ("current_stage", "Current Stage", 16),
    ("next_action", "Next Action", 28),
    ("next_action_due_date", "Next Action Due Date", 14),
    ("campaign_id", "Campaign ID", 26),
    ("sector", "Sector", 20),
    ("geography", "Geography", 14),
    ("area", "Area", 22),
    ("priority", "Priority", 9),
    ("opportunity_type", "Opportunity Type", 14),
    ("accessibility", "Accessibility", 14),
    ("role", "Role", 22),
    ("contact_phone", "Contact Phone", 15),
    ("decision_maker_linkedin", "Decision-Maker LinkedIn", 26),
    ("website", "Website", 26),
    ("ready_to_email", "Ready To Email", 12),
    ("outreach_rank", "Outreach Rank", 11),
    ("do_not_contact_manual", "Do Not Contact (Manual)", 16),
    ("notes", "Notes", 34),
    ("research_pipeline_stage", "Research Pipeline Stage", 16),
    ("withheld_at_outreach", "Withheld At Outreach", 14),
    ("last_activity_type", "Last Activity Type", 22),
    ("last_activity_date", "Last Activity Date", 13),
    ("blocked", "Blocked", 9),
    ("cadence_next_action_raw", "Cadence Next Action (raw)", 26),
    ("cadence_days_raw", "Cadence Days (raw)", 12),
    ("cadence_stage_raw", "Cadence Stage (raw)", 14),
    ("raw_due_date", "Raw Due Date", 13),
    ("overdue", "Overdue", 9),
    ("due_today", "Due Today", 9),
]
PI = {key: i for i, (key, _label, _w) in enumerate(PROSPECTS_COLUMNS)}  # 0-indexed
HELPER_COL_START = PI["research_pipeline_stage"]  # first column of the collapsible helper block

ACTIVITIES_COLUMNS = [
    # "Row #", not "Activity ID": a =ROW()-based position, not a stable
    # identifier - it renumbers if rows are sorted or one is deleted.
    # Nothing in this codebase treats it as an identity; see the note
    # above the formula loop in build_activities_sheet().
    ("activity_id", "Row #", 8),
    ("prospect_id", "Prospect ID", 34),
    ("business", "Business", 26),
    ("activity_type", "Activity Type", 26),
    ("activity_date", "Activity Date", 13),
    ("amount", "Amount (GBP)", 12),
    ("plan", "Plan", 10),
    ("notes", "Notes", 40),
    ("logged_by", "Logged By", 10),
    ("logged_at", "Logged At", 18),
]
AI = {key: i for i, (key, _label, _w) in enumerate(ACTIVITIES_COLUMNS)}

CADENCE_COLUMNS = [
    ("key", "Key", 26),
    ("label", "Activity Type", 26),
    ("next_action_label", "Next Action", 30),
    ("cadence_days", "Cadence Days", 12),
    ("stage_label", "Stage", 14),
    ("stops_cold_followup", "Stops Cold Follow-up", 14),
    ("blocks_outreach", "Blocks Outreach", 13),
    ("is_revenue_event", "Revenue Event", 12),
]
CI = {key: i for i, (key, _label, _w) in enumerate(CADENCE_COLUMNS)}

# Fixed sheet layout - deliberately NOT derived from add_table()'s return
# values. Cross-sheet formulas below use plain absolute ranges rather than
# Table structured references: a structured self-reference ([@[Column]])
# combined with a reference to another sheet in the SAME formula corrupts
# the file for Excel's automation loader (reproduced and confirmed while
# building this workbook - each half works fine alone, only the
# combination breaks; openpyxl reads the result but real Excel refuses to
# open it). Plain cell/range references sidestep the bug entirely and are
# the same pattern tools/prospect-compiler/build_workbook.py already uses.
# Capacities are generous headroom so rows added by hand in Excel between
# rebuilds stay inside the formula ranges without needing a rebuild.
ACTIVITIES_HEADER_ROW = 3   # 1-indexed Excel row
ACTIVITIES_DATA_FIRST = 4
ACTIVITIES_CAPACITY = 500
ACTIVITIES_DATA_LAST = ACTIVITIES_DATA_FIRST + ACTIVITIES_CAPACITY - 1

SETTINGS_CADENCE_HEADER_ROW = 7
SETTINGS_CADENCE_DATA_FIRST = 8
CADENCE_CAPACITY = 60
SETTINGS_CADENCE_DATA_LAST = SETTINGS_CADENCE_DATA_FIRST + CADENCE_CAPACITY - 1

PIPELINE_STAGES = [
    "Research complete", "Qualified", "Outreach ready", "Prepared",
    "Contacted", "Follow-up due", "Replied", "Meeting", "Audit",
    "Foundation", "Ongoing", "Won", "Lost", "Do Not Contact",
]
WON_STAGES = {"Audit", "Foundation", "Ongoing"}


EXCEL_EPOCH = date(1899, 12, 30)  # standard Excel serial-date epoch


def cadence_by_label(table=None):
    """cadence.cadence_by_key() indexes by the internal machine key
    (e.g. 'OUTREACH_PREPARED') - correct for cadence.py's own tests, but
    every activity_type value this workbook actually stores (Activities
    rows, the Excel dropdown, the CadenceRules join column everything
    keys off) is the human-readable label ('Outreach prepared'), to match
    what the owner picks from the dropdown. Any Python-side lookup here
    must be indexed the same way Excel's own MATCH()/INDEX() formulas are,
    or the cached value silently falls through to blank while Excel's
    live formula computes the right answer - reproduced and fixed while
    building this workbook."""
    table = table if table is not None else cad.default_cadence_table()
    return {row["label"]: row for row in table}


def excel_serial(d):
    """write_formula()'s cached-value argument does not convert a Python
    date the way write_datetime() does - passed directly, xlsxwriter writes
    its literal ISO string into an untyped (implicitly numeric) <v> element,
    which is a genuine OOXML schema violation: Excel's own loader refuses
    to open the file outright rather than repairing it (reproduced and
    confirmed while building this workbook - openpyxl reads it anyway,
    which is what made this easy to miss). Every cached value for a
    date-returning formula must go through this conversion first."""
    return (d - EXCEL_EPOCH).days


def col_letter(idx0):
    letters = ""
    n = idx0 + 1
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def act_range(key):
    c = col_letter(AI[key])
    return f"Activities!${c}${ACTIVITIES_DATA_FIRST}:${c}${ACTIVITIES_DATA_LAST}"


PROSPECTS_DATA_FIRST = 4  # 1-indexed; matches build_prospects_sheet's first_row=2 (0idx) + 1 header row.


def prospects_range(key, n_prospects):
    """Absolute range on Prospects for a given column, sized to the actual
    prospect count at this build (Prospects itself is fully regenerated
    every build_crm.py run, unlike Activities, so no extra headroom is
    needed here - callers writing formulas that must survive until the
    next rebuild, e.g. Activities' Business lookup, still get a fixed
    absolute range rather than a structured reference, for the same
    reason documented above ACTIVITIES_HEADER_ROW."""
    c = col_letter(PI[key])
    last = PROSPECTS_DATA_FIRST + max(n_prospects, 1) - 1
    return f"Prospects!${c}${PROSPECTS_DATA_FIRST}:${c}${last}"


def cadence_range(key):
    c = col_letter(CI[key])
    return f"Settings!${c}${SETTINGS_CADENCE_DATA_FIRST}:${c}${SETTINGS_CADENCE_DATA_LAST}"


# ---------------------------------------------------------------------------
# Load tracker.json
# ---------------------------------------------------------------------------

def load_tracker(tracker_path):
    with open(tracker_path, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Load the prior workbook, if any - the only source of manual truth.
# ---------------------------------------------------------------------------

def _table_rows(ws, table_name):
    """(header_row_cells, data_rows_as_value_tuples) for a named Excel
    Table on ws, located via the table's own declared range rather than an
    assumed row 1 - every sheet in this workbook carries a title/subtitle
    above its table, so the header is never actually on row 1."""
    tbl = ws.tables[table_name]
    cell_range = ws[tbl.ref]
    rows = list(cell_range)
    header_row = rows[0]
    data_rows = [tuple(c.value for c in r) for r in rows[1:]]
    return header_row, data_rows


def load_existing_workbook(path):
    """Returns (activities, manual_fields). activities: list of dicts, one
    per Activities row already logged. manual_fields: {prospect_id: {
    'do_not_contact_manual': bool, 'notes': str}}. Never touches
    tracker.json or anything under wardith-runs/<slug>/ - read-only against
    the previous CRM workbook only."""
    if not path.exists():
        return [], {}
    wb = load_workbook(path, data_only=False)
    activities = []
    if "Activities" in wb.sheetnames and "Activities" in wb["Activities"].tables:
        ws = wb["Activities"]
        header_row, data_rows = _table_rows(ws, "Activities")
        headers = [c.value for c in header_row]
        idx = {h: i for i, h in enumerate(headers) if h}
        for row in data_rows:
            if not row or not row[idx.get("Prospect ID", 1)]:
                continue
            pid = row[idx["Prospect ID"]]
            atype = row[idx.get("Activity Type")]
            adate = row[idx.get("Activity Date")]
            if not pid or not atype or not adate:
                continue
            if hasattr(adate, "date"):
                adate = adate.date()
            elif isinstance(adate, str):
                adate = parse_date(adate)
            if not isinstance(adate, date):
                continue
            activities.append({
                "prospect_id": pid,
                "activity_type": atype,
                "activity_date": adate,
                "amount": row[idx.get("Amount (GBP)")] if "Amount (GBP)" in idx else None,
                "plan": row[idx.get("Plan")] if "Plan" in idx else None,
                "notes": row[idx.get("Notes")] if "Notes" in idx else None,
                "logged_by": row[idx.get("Logged By")] if "Logged By" in idx else "Owner",
            })
    manual_fields = {}
    if "Prospects" in wb.sheetnames and "Prospects" in wb["Prospects"].tables:
        ws = wb["Prospects"]
        header_row, data_rows = _table_rows(ws, "Prospects")
        headers = [c.value for c in header_row]
        idx = {h: i for i, h in enumerate(headers) if h}
        if "Prospect ID" in idx:
            for row in data_rows:
                if not row or not row[idx["Prospect ID"]]:
                    continue
                pid = row[idx["Prospect ID"]]
                manual_fields[pid] = {
                    "do_not_contact_manual": bool(row[idx["Do Not Contact (Manual)"]]) if "Do Not Contact (Manual)" in idx else False,
                    "notes": row[idx["Notes"]] if "Notes" in idx and row[idx["Notes"]] else "",
                }
    wb.close()
    return activities, manual_fields


def merge_activities(existing_activities, tracker):
    """existing_activities (from the prior workbook) plus one auto-logged
    'Outreach prepared' row for every prospect the (bug-fixed) importer
    marked genuinely PREPARED, if that fact isn't already logged. Import
    never invents SENT/REPLIED/etc. - those stay human-only - but a
    genuinely drafted, non-withheld outreach-prep record is itself a
    system-known fact, safe to log automatically and idempotently."""
    activities = list(existing_activities)
    already_logged = {
        (a["prospect_id"], a["activity_type"])
        for a in activities
    }
    prepared_label = cad.cadence_by_key()["OUTREACH_PREPARED"]["label"]
    next_id = (max((a.get("activity_id", 0) or 0) for a in activities), 0)[0] + 1 if activities else 1
    added = 0
    for prospect_id, p in sorted(tracker.get("prospects", {}).items()):
        activity = p.get("activity", {})
        if activity.get("outreach_status") != "PREPARED":
            continue
        if (prospect_id, prepared_label) in already_logged:
            continue
        prepared_date = parse_date(activity.get("prepared_date"))
        if not prepared_date:
            continue
        activities.append({
            "prospect_id": prospect_id,
            "activity_type": prepared_label,
            "activity_date": prepared_date,
            "amount": None, "plan": None,
            "notes": "Auto-logged from a genuinely drafted, non-withheld /outreach record.",
            "logged_by": "Import",
        })
        added += 1
    for i, a in enumerate(activities, start=1):
        a["activity_id"] = i
    return activities, added


# ---------------------------------------------------------------------------
# Build the per-prospect row model (python-side, mirrors every Excel formula)
# ---------------------------------------------------------------------------

def build_prospect_rows(tracker, activities, manual_fields, today):
    cadence_map = cadence_by_label()
    activities_by_prospect = {}
    for a in activities:
        activities_by_prospect.setdefault(a["prospect_id"], []).append(a)

    campaigns = tracker.get("campaigns", {})
    rows = []
    for prospect_id, p in sorted(tracker.get("prospects", {}).items()):
        research = p.get("research", {}) or {}
        campaign = campaigns.get(research.get("campaign_id"), {})
        manual = manual_fields.get(prospect_id, {})
        manual_dnc = bool(manual.get("do_not_contact_manual", False))
        acts = activities_by_prospect.get(prospect_id, [])

        research_for_cadence = {
            "pipeline_stage": campaign.get("pipeline_stage"),
            "ready_to_email": research.get("ready_to_email"),
            "withheld_at_outreach": research.get("withheld_at_outreach"),
        }
        state = cad.compute_prospect_state(research_for_cadence, acts, manual_do_not_contact=manual_dnc,
                                            cadence=cadence_map, today=today)

        rule = cadence_map.get(state["last_activity_type"], {}) if state["last_activity_type"] else {}
        cadence_days_raw = rule.get("cadence_days")
        raw_due = None
        if state["last_activity_date"] and cadence_days_raw is not None:
            raw_due = state["last_activity_date"] + timedelta(days=cadence_days_raw)

        row = {
            "prospect_id": prospect_id,
            "business": research.get("business") or "",
            "contact_person": research.get("contact_person") or "",
            "contact_email": research.get("contact_email") or "",
            "current_stage": state["stage"],
            "next_action": state["next_action"],
            "next_action_due_date": state["next_action_due_date"],
            "campaign_id": research.get("campaign_id") or "",
            "sector": campaign.get("sector") or "",
            "geography": campaign.get("geography") or "",
            "area": research.get("area") or "",
            "priority": research.get("priority") or "",
            "opportunity_type": research.get("opportunity_type") or "",
            "accessibility": research.get("accessibility") or "",
            "role": research.get("role") or "",
            "contact_phone": research.get("contact_phone") or "",
            "decision_maker_linkedin": research.get("decision_maker_linkedin") or "",
            "website": research.get("website") or "",
            "ready_to_email": research.get("ready_to_email") or "",
            "outreach_rank": research.get("outreach_rank"),
            "do_not_contact_manual": manual_dnc,
            "notes": manual.get("notes", "") or "",
            "research_pipeline_stage": campaign.get("pipeline_stage") or "",
            "withheld_at_outreach": bool(research.get("withheld_at_outreach")),
            "last_activity_type": state["last_activity_type"] or "",
            "last_activity_date": state["last_activity_date"],
            "blocked": state["blocked"],
            "cadence_next_action_raw": rule.get("next_action_label", "") if state["last_activity_type"] else "",
            "cadence_days_raw": cadence_days_raw if state["last_activity_type"] else None,
            "cadence_stage_raw": rule.get("stage_label", "") if state["last_activity_type"] else "",
            "raw_due_date": raw_due,
            "overdue": state["overdue"],
            "due_today": state["next_action_due_date"] == today if state["next_action_due_date"] else False,
        }
        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Formats
# ---------------------------------------------------------------------------

def make_formats(wb):
    return {
        "title": wb.add_format({"bold": True, "font_size": 16, "font_color": NAVY}),
        "subtitle": wb.add_format({"font_size": 10, "italic": True, "font_color": INK}),
        "h2": wb.add_format({"bold": True, "font_size": 12, "font_color": NAVY}),
        "h3": wb.add_format({"bold": True, "font_size": 10.5, "font_color": NAVY}),
        "body": wb.add_format({"font_size": 10, "text_wrap": True, "valign": "top"}),
        "body_small": wb.add_format({"font_size": 9.5, "text_wrap": True, "valign": "top"}),
        "header": wb.add_format({"bold": True, "font_color": PAPER, "bg_color": NAVY, "text_wrap": True,
                                  "valign": "vcenter", "border": 1, "border_color": RULE}),
        "cell": wb.add_format({"valign": "top", "border": 1, "border_color": RULE}),
        "cell_wrap": wb.add_format({"valign": "top", "border": 1, "border_color": RULE, "text_wrap": True}),
        "cell_bold": wb.add_format({"valign": "top", "border": 1, "border_color": RULE, "bold": True}),
        "cell_date": wb.add_format({"valign": "top", "border": 1, "border_color": RULE, "num_format": "yyyy-mm-dd"}),
        "cell_money": wb.add_format({"valign": "top", "border": 1, "border_color": RULE, "num_format": "£#,##0"}),
        "cell_center": wb.add_format({"valign": "top", "border": 1, "border_color": RULE, "align": "center"}),
        "input_cell": wb.add_format({"valign": "top", "border": 1, "border_color": BRASS, "bg_color": "#FBF6EA"}),
        "input_date": wb.add_format({"valign": "top", "border": 1, "border_color": BRASS, "bg_color": "#FBF6EA",
                                      "num_format": "yyyy-mm-dd"}),
        "kpi_num": wb.add_format({"bold": True, "font_size": 20, "font_color": NAVY, "align": "center"}),
        "kpi_label": wb.add_format({"font_size": 9.5, "align": "center", "font_color": INK}),
    }


def write_table_header(ws, row0, col0, columns, fmts):
    for i, (_key, label, width) in enumerate(columns):
        ws.write(row0, col0 + i, label, fmts["header"])
        ws.set_column(col0 + i, col0 + i, width)


def add_table(ws, first_row, first_col, n_data_rows, columns, name, style="Table Style Medium 9"):
    """Column widths + an Excel Table over [first_row..first_row+n_data_rows],
    header row written by the table itself. Data rows must already be
    written (or left blank for the user to fill in) before this call."""
    for i, (_key, _label, width) in enumerate(columns):
        ws.set_column(first_col + i, first_col + i, width)
    last_row = first_row + max(n_data_rows, 1)
    last_col = first_col + len(columns) - 1
    ws.add_table(first_row, first_col, last_row, last_col, {
        "name": name,
        "banded_rows": True,
        "style": style,
        "columns": [{"header": label} for _key, label, _w in columns],
    })
    return last_row, last_col


# ---------------------------------------------------------------------------
# Settings sheet
# ---------------------------------------------------------------------------

def build_settings_sheet(wb, fmts, cadence_table):
    ws = wb.add_worksheet("Settings")
    ws.hide_gridlines(2)
    ws.freeze_panes(3, 0)
    ws.write(0, 0, "Settings", fmts["title"])
    ws.write(1, 0, "Every rule the workbook's formulas key off, in one place. Edit a value here and every "
                    "Next Action / Due Date in Prospects updates - no formula or code change needed.",
             fmts["subtitle"])

    r = 3
    ws.write(r, 0, "Cadence rules", fmts["h2"])
    r += 1
    ws.write(r, 0, "One row per activity type. \"Activity Type\" is the exact text used in the Activities "
                    "sheet's dropdown and in every lookup formula - renaming a label here renames it "
                    "everywhere; do not leave two rows with the same text.", fmts["body_small"])
    r += 2
    cadence_first_row = r
    assert cadence_first_row == SETTINGS_CADENCE_HEADER_ROW - 1, (
        "Settings sheet layout drifted - update SETTINGS_CADENCE_HEADER_ROW to match, "
        "every cross-sheet formula in build_prospects_sheet depends on this row being fixed."
    )
    for i, row in enumerate(cadence_table):
        rr = cadence_first_row + 1 + i
        ws.write(rr, 0, row["key"], fmts["cell"])
        ws.write(rr, 1, row["label"], fmts["cell_bold"])
        ws.write(rr, 2, row["next_action_label"], fmts["cell"])
        if row["cadence_days"] is None:
            ws.write_blank(rr, 3, None, fmts["cell_center"])
        else:
            ws.write_number(rr, 3, row["cadence_days"], fmts["cell_center"])
        ws.write(rr, 4, row["stage_label"], fmts["cell"])
        ws.write_boolean(rr, 5, row["stops_cold_followup"], fmts["cell_center"])
        ws.write_boolean(rr, 6, row["blocks_outreach"], fmts["cell_center"])
        ws.write_boolean(rr, 7, row["is_revenue_event"], fmts["cell_center"])
    cadence_last_row, _ = add_table(ws, cadence_first_row, 0, CADENCE_CAPACITY, CADENCE_COLUMNS, "CadenceRules",
                                     style="Table Style Medium 9")
    ws.set_row(cadence_first_row, 30)

    r = cadence_last_row + 3
    ws.write(r, 0, "Recurring-plan monthly prices", fmts["h2"])
    r += 1
    ws.write(r, 0, "playbook/services.md - the published prices. Edit here if they change; Revenue sheet "
                    "reads amounts logged on Activities directly, this table is reference only.",
             fmts["body_small"])
    r += 2
    plan_first_row = r
    ws.write(r, 0, "Plan", fmts["header"])
    ws.write(r, 1, "Monthly Price (GBP)", fmts["header"])
    ws.set_column(0, 0, 26)
    ws.set_column(1, 1, 16)
    for i, (plan, price) in enumerate(PLAN_PRICES):
        ws.write(plan_first_row + 1 + i, 0, plan, fmts["cell"])
        ws.write_number(plan_first_row + 1 + i, 1, price, fmts["cell_money"])
    ws.add_table(plan_first_row, 0, plan_first_row + len(PLAN_PRICES), 1, {
        "name": "PlanPrices", "style": "Table Style Medium 9",
        "columns": [{"header": "Plan"}, {"header": "Monthly Price (GBP)"}],
    })

    r = plan_first_row + len(PLAN_PRICES) + 3
    ws.write(r, 0, "One-off prices", fmts["h2"])
    r += 1
    ws.write(r, 0, f"Audit: £{AUDIT_PRICE} one-off. Foundation: £{FOUNDATION_PRICE} one-off. "
                    f"playbook/services.md.", fmts["body_small"])
    return ws, cadence_first_row, cadence_last_row


# ---------------------------------------------------------------------------
# Activities sheet - append-only interaction history
# ---------------------------------------------------------------------------

def build_activities_sheet(wb, fmts, activities, business_by_prospect_id, n_prospects):
    ws = wb.add_worksheet("Activities")
    ws.hide_gridlines(2)
    ws.write(0, 0, "Activities - append-only. Add a row: pick Prospect ID, pick Activity Type, set the "
                    "date - Business fills in on its own. Next Action / Due Date on Prospects recalculate "
                    "automatically.", fmts["subtitle"])
    ws.set_row(0, 16)
    first_row = 2
    assert first_row == ACTIVITIES_HEADER_ROW - 1, (
        "Activities sheet layout drifted - update ACTIVITIES_HEADER_ROW to match, every "
        "cross-sheet formula in build_prospects_sheet depends on this row being fixed."
    )
    ws.freeze_panes(first_row + 1, 0)

    activities_sorted = sorted(activities, key=lambda a: (a["prospect_id"], a["activity_date"]))
    for i, a in enumerate(activities_sorted):
        rr = first_row + 1 + i
        fmt = fmts["cell"]
        ws.write(rr, AI["prospect_id"], a["prospect_id"], fmt)
        # Business (see the formula loop below, after add_table): a live
        # lookup keyed off this row's own Prospect ID, not a static value -
        # the user only ever needs to pick Prospect ID, Activity Type and
        # Activity Date; Business fills in on its own and stays correct if
        # a prospect's name is ever corrected on a later rebuild.
        ws.write(rr, AI["activity_type"], a["activity_type"], fmt)
        ws.write_datetime(rr, AI["activity_date"], a["activity_date"], fmts["cell_date"])
        if a.get("amount") not in (None, ""):
            ws.write_number(rr, AI["amount"], float(a["amount"]), fmts["cell_money"])
        else:
            ws.write_blank(rr, AI["amount"], None, fmts["input_cell"])
        ws.write(rr, AI["plan"], a.get("plan") or "", fmts["input_cell"] if not a.get("plan") else fmt)
        ws.write(rr, AI["notes"], a.get("notes") or "", fmts["cell_wrap"])
        ws.write(rr, AI["logged_by"], a.get("logged_by") or "Owner", fmt)
        ws.write(rr, AI["logged_at"], a.get("logged_at") or "", fmt)

    # Reserve blank rows up to ACTIVITIES_CAPACITY: build_prospects_sheet's
    # formulas reference this sheet by a fixed absolute range (see the
    # module-level layout comment), so headroom here is what lets a row
    # added by hand in Excel between rebuilds be picked up immediately.
    if len(activities_sorted) > ACTIVITIES_CAPACITY:
        raise ValueError(f"{len(activities_sorted)} activities exceeds ACTIVITIES_CAPACITY="
                          f"{ACTIVITIES_CAPACITY} - raise the constant and rebuild.")
    last_row, last_col = add_table(ws, first_row, 0, ACTIVITIES_CAPACITY, ACTIVITIES_COLUMNS, "Activities",
                                    style="Table Style Medium 9")

    # Row # and Business: both formulas for every reserved row, including
    # ones the user adds by hand, so they're already live the moment a
    # Prospect ID is picked.
    #
    # Row # (=ROW()-N) is a POSITION, not an identity - it renumbers if
    # rows are ever sorted or one is deleted. Nothing in this codebase
    # treats it as a stable id: the one place a row needs matching
    # (merge_activities' "already logged?" check) keys off
    # (prospect_id, activity_type), never this column. If an activity ever
    # needs a real identifier, add one deliberately - don't repurpose this.
    business_cache_by_row = {}
    for i, a in enumerate(activities_sorted):
        business_cache_by_row[first_row + 1 + i] = business_by_prospect_id.get(a["prospect_id"], "")
    pid_col = col_letter(AI["prospect_id"])
    pros_id_range = prospects_range("prospect_id", n_prospects)
    pros_biz_range = prospects_range("business", n_prospects)
    for rr in range(first_row + 1, last_row + 1):
        xr = rr + 1
        ws.write_formula(rr, AI["activity_id"], f"=ROW()-{first_row + 1}", fmts["cell_center"], rr - first_row)
        f_business = f'=IFERROR(INDEX({pros_biz_range},MATCH({pid_col}{xr},{pros_id_range},0)),"")'
        ws.write_formula(rr, AI["business"], f_business, fmts["cell"], business_cache_by_row.get(rr, ""))

    # A data-validation list formula that points at a Table on a DIFFERENT
    # sheet must go through a workbook-level defined name - Excel's
    # automation layer (unlike interactive Excel with a repair prompt)
    # refuses to open a file with a raw cross-sheet structured reference as
    # a validation source, failing outright rather than repairing.
    ws.data_validation(first_row + 1, AI["prospect_id"], last_row, AI["prospect_id"],
                        {"validate": "list", "source": "=ProspectIDList"})
    ws.data_validation(first_row + 1, AI["activity_type"], last_row, AI["activity_type"],
                        {"validate": "list", "source": "=ActivityTypeList"})
    ws.data_validation(first_row + 1, AI["plan"], last_row, AI["plan"],
                        {"validate": "list", "source": ["", "Maintain", "Grow", "Lead"]})
    ws.data_validation(first_row + 1, AI["logged_by"], last_row, AI["logged_by"],
                        {"validate": "list", "source": ["Owner", "Import"]})
    return ws, first_row, last_row


TODAY_SLICE = ["business", "contact_person", "contact_email", "current_stage", "next_action", "next_action_due_date"]


TODAY_BLOCK_SPILL_CAPACITY = 40  # a practical ceiling per block - see below for how overflow past it is
# handled, not just deferred. Raise this if a block is routinely showing the overflow message.


def _today_block(ws, fmts, start_row, title, criteria_formula, sort_col_1based, rows_matching):
    ws.write(start_row, 0, title, fmts["h2"])
    anchor_row = start_row + 2
    headers = ["Business", "Contact Person", "Contact Email", "Current Stage", "Next Action", "Due Date"]
    for i, h in enumerate(headers):
        ws.write(anchor_row, i, h, fmts["header"])
    # A dynamic-array spill only takes explicit formatting from its own
    # write call, which covers just the anchor cell here (the spill size
    # isn't known at build time) - spilled cells beyond the anchor keep
    # whatever format was already sitting in that cell, so the Due Date
    # column's date format has to be pre-set across the whole reserved
    # capacity before the formula is written, or it displays as a raw
    # serial number once real dates spill into it.
    date_col = len(headers) - 1
    for r in range(anchor_row + 1, anchor_row + 1 + TODAY_BLOCK_SPILL_CAPACITY):
        ws.write_blank(r, date_col, None, fmts["cell_date"])

    # The live match count can legitimately exceed what was true at the
    # last rebuild (Excel's own TODAY() at open time can differ) - and
    # must never be silently hidden. Two things enforce that:
    #  1. TAKE(...) hard-caps the spill at TODAY_BLOCK_SPILL_CAPACITY rows,
    #     so a genuinely larger live result can never overflow into the
    #     next block and break it with #SPILL! (reproduced and fixed while
    #     building this workbook - sizing the reserved space off the
    #     build-time count risked exactly that).
    #  2. The count line below is itself a live formula (not the
    #     build-time count) and switches to an explicit overflow message
    #     the moment the true total exceeds the cap.
    # --(...) forces explicit boolean-to-number coercion: reproduced and
    # confirmed while building this workbook - a bare SUMPRODUCT of a
    # single-comparison boolean array against a structured Table reference
    # (e.g. Prospects[Due Today]=TRUE) silently returns 0 instead of the
    # real count in this Excel installation, while the double-unary form
    # and COUNTIF both compute correctly. Always coerce; never rely on
    # SUMPRODUCT's implicit boolean handling here.
    count_expr = f'SUMPRODUCT(--({criteria_formula}))'
    count_formula = (f'=IF({count_expr}>{TODAY_BLOCK_SPILL_CAPACITY},'
                      f'"Showing {TODAY_BLOCK_SPILL_CAPACITY} of "&{count_expr}&'
                      f'" - open Prospects for the remainder.",{count_expr}&" matching - recalculates live.")')
    live_count = len(rows_matching)
    cached_count_text = (f"Showing {TODAY_BLOCK_SPILL_CAPACITY} of {live_count} - open Prospects for the "
                          f"remainder." if live_count > TODAY_BLOCK_SPILL_CAPACITY
                          else f"{live_count} matching - recalculates live.")
    ws.write_formula(start_row + 1, 0, count_formula, fmts["body_small"], cached_count_text)

    formula = (f'=IFERROR(TAKE(SORT(FILTER(Prospects[[Business]:[Next Action Due Date]],{criteria_formula}),'
               f'{sort_col_1based},1),{TODAY_BLOCK_SPILL_CAPACITY}),"None.")')
    first_val = rows_matching[0]["business"] if rows_matching else "None."
    ws.write_dynamic_array_formula(anchor_row + 1, 0, anchor_row + 1, 0, formula, fmts["cell"], first_val)
    return anchor_row + 2 + TODAY_BLOCK_SPILL_CAPACITY


def build_today_sheet(wb, fmts, rows, today):
    ws = wb.add_worksheet("Today")
    ws.hide_gridlines(2)
    ws.set_column(0, 5, 24)
    ws.write(0, 0, f"Today - {today.isoformat()}", fmts["title"])
    ws.write(1, 0, "The daily home screen. Everything below is live: it recalculates the moment an "
                    "Activities row is added or a Settings cadence value changes.", fmts["subtitle"])
    r = 3

    overdue_rows = sorted((row for row in rows if row["overdue"] and not row["blocked"]),
                           key=lambda row: row["next_action_due_date"])
    r = _today_block(ws, fmts, r, "Overdue actions", "Prospects[Overdue]=TRUE", 6, overdue_rows)

    due_today_rows = [row for row in rows if row["due_today"] and not row["blocked"]]
    r = _today_block(ws, fmts, r, "Due today", "Prospects[Due Today]=TRUE", 1, due_today_rows)

    in7 = today + timedelta(days=7)
    upcoming_rows = sorted(
        (row for row in rows if row["next_action_due_date"] and not row["blocked"] and not row["overdue"]
         and not row["due_today"] and today < row["next_action_due_date"] <= in7),
        key=lambda row: row["next_action_due_date"])
    upcoming_criteria = ('(Prospects[Next Action Due Date]<>"")*(Prospects[Next Action Due Date]>TODAY())*'
                          f'(Prospects[Next Action Due Date]<=TODAY()+7)*(Prospects[Blocked]<>TRUE)')
    r = _today_block(ws, fmts, r, "Upcoming (next 7 days)", upcoming_criteria, 6, upcoming_rows)

    ready_rows = [row for row in rows if row["current_stage"] == "Outreach ready"]
    r = _today_block(ws, fmts, r, "New outreach-ready prospects", 'Prospects[Current Stage]="Outreach ready"',
                      1, ready_rows)

    replied_rows = [row for row in rows if row["current_stage"] == "Replied"]
    r = _today_block(ws, fmts, r, "Replies requiring action", 'Prospects[Current Stage]="Replied"', 1, replied_rows)

    return ws


def build_campaigns_sheet(wb, fmts, tracker, rows):
    ws = wb.add_worksheet("Campaigns")
    ws.hide_gridlines(2)
    ws.write(0, 0, "Campaigns - refreshed from tracker.json on every rebuild.", fmts["subtitle"])
    first_row = 2
    columns = [
        ("campaign_id", "Campaign ID", 30), ("sector", "Sector", 22), ("geography", "Geography", 14),
        ("run_date", "Run Date", 12), ("pipeline_stage", "Pipeline Stage", 16),
        ("market_count", "Market Count", 12), ("outreach_count", "Outreach Count", 12),
        ("excluded_count", "Excluded Count", 12), ("prospects_in_crm", "Prospects In CRM", 13),
        ("prepared", "Prepared", 10), ("contacted", "Contacted", 10), ("replied", "Replied", 10),
        ("won", "Won", 8), ("lost", "Lost", 8),
    ]
    campaigns = tracker.get("campaigns", {})
    for i, (campaign_id, c) in enumerate(sorted(campaigns.items())):
        rr = first_row + 1 + i
        crm_rows = [row for row in rows if row["campaign_id"] == campaign_id]
        ws.write(rr, 0, campaign_id, fmts["cell_bold"])
        ws.write(rr, 1, c.get("sector") or "", fmts["cell"])
        ws.write(rr, 2, c.get("geography") or "", fmts["cell"])
        ws.write(rr, 3, c.get("run_date") or "", fmts["cell"])
        ws.write(rr, 4, c.get("pipeline_stage") or "", fmts["cell"])
        ws.write(rr, 5, c.get("market_count") if c.get("market_count") is not None else "", fmts["cell_center"])
        ws.write(rr, 6, c.get("outreach_count") if c.get("outreach_count") is not None else "", fmts["cell_center"])
        ws.write(rr, 7, c.get("excluded_count") if c.get("excluded_count") is not None else "", fmts["cell_center"])
        ws.write_number(rr, 8, len(crm_rows), fmts["cell_center"])
        ws.write_number(rr, 9, sum(1 for row in crm_rows if row["current_stage"] == "Prepared"), fmts["cell_center"])
        ws.write_number(rr, 10, sum(1 for row in crm_rows if row["current_stage"] in ("Contacted", "Follow-up due")),
                         fmts["cell_center"])
        ws.write_number(rr, 11, sum(1 for row in crm_rows if row["current_stage"] == "Replied"), fmts["cell_center"])
        ws.write_number(rr, 12, sum(1 for row in crm_rows if row["current_stage"] in WON_STAGES), fmts["cell_center"])
        ws.write_number(rr, 13, sum(1 for row in crm_rows if row["current_stage"] == "Lost"), fmts["cell_center"])
    add_table(ws, first_row, 0, len(campaigns), columns, "Campaigns", style="Table Style Medium 9")
    return ws


def build_pipeline_sheet(wb, fmts, rows, tracker):
    ws = wb.add_worksheet("Pipeline")
    ws.hide_gridlines(2)
    ws.write(0, 0, "Pipeline", fmts["title"])
    ws.write(1, 0, "The funnel across every campaign in the CRM.", fmts["subtitle"])
    r = 3
    ws.write(r, 0, "Stage", fmts["header"])
    ws.write(r, 1, "Count", fmts["header"])
    ws.set_column(0, 0, 22)
    ws.set_column(1, 1, 10)
    counts = {}
    for stage in PIPELINE_STAGES:
        if stage == "Won":
            counts[stage] = sum(1 for row in rows if row["current_stage"] in WON_STAGES)
        else:
            counts[stage] = sum(1 for row in rows if row["current_stage"] == stage)
    max_count = max(counts.values()) if counts else 1
    for i, stage in enumerate(PIPELINE_STAGES):
        rr = r + 1 + i
        ws.write(rr, 0, stage, fmts["cell_bold"])
        ws.write_number(rr, 1, counts[stage], fmts["cell_center"])
    ws.conditional_format(r + 1, 1, r + len(PIPELINE_STAGES), 1, {
        "type": "data_bar", "bar_color": NAVY, "bar_solid": True,
    })

    r2 = r + len(PIPELINE_STAGES) + 3
    ws.write(r2, 0, "By campaign", fmts["h2"])
    r2 += 1
    campaigns = sorted(tracker.get("campaigns", {}).keys())
    matrix_stages = ["Outreach ready", "Prepared", "Contacted", "Follow-up due", "Replied", "Meeting", "Won", "Lost"]
    ws.write(r2, 0, "Campaign", fmts["header"])
    for j, stage in enumerate(matrix_stages):
        ws.write(r2, 1 + j, stage, fmts["header"])
        ws.set_column(1 + j, 1 + j, 13)
    for i, campaign_id in enumerate(campaigns):
        rr = r2 + 1 + i
        ws.write(rr, 0, campaign_id, fmts["cell"])
        crm_rows = [row for row in rows if row["campaign_id"] == campaign_id]
        for j, stage in enumerate(matrix_stages):
            if stage == "Won":
                n = sum(1 for row in crm_rows if row["current_stage"] in WON_STAGES)
            else:
                n = sum(1 for row in crm_rows if row["current_stage"] == stage)
            ws.write_number(rr, 1 + j, n, fmts["cell_center"])
    return ws


def build_revenue_sheet(wb, fmts, activities, business_by_prospect_id, cadence_map):
    ws = wb.add_worksheet("Revenue")
    ws.hide_gridlines(2)
    ws.write(0, 0, "Revenue", fmts["title"])
    ws.write(1, 0, "Audit, Foundation and ongoing-plan revenue, from Activities rows flagged as a revenue "
                    "event in Settings. Ongoing-plan revenue is recognised once, in the month it starts - "
                    "this sheet does not project it forward month by month.", fmts["subtitle"])
    r = 3
    revenue_events = [a for a in activities if cadence_map.get(a["activity_type"], {}).get("is_revenue_event")
                       or a.get("amount")]
    columns = [
        ("business", "Business", 28), ("activity_type", "Activity Type", 20), ("activity_date", "Date", 12),
        ("amount", "Amount (GBP)", 13), ("plan", "Plan", 10), ("month", "Month", 10),
    ]
    first_data = r + 1
    monthly = {}
    audit_total = foundation_total = ongoing_total = 0.0
    for i, a in enumerate(sorted(revenue_events, key=lambda a: a["activity_date"])):
        rr = first_data + i
        amount = float(a.get("amount") or 0)
        month = a["activity_date"].strftime("%Y-%m")
        ws.write(rr, 0, business_by_prospect_id.get(a["prospect_id"], ""), fmts["cell"])
        ws.write(rr, 1, a["activity_type"], fmts["cell"])
        ws.write_datetime(rr, 2, a["activity_date"], fmts["cell_date"])
        ws.write_number(rr, 3, amount, fmts["cell_money"])
        ws.write(rr, 4, a.get("plan") or "", fmts["cell"])
        ws.write(rr, 5, month, fmts["cell_center"])
        monthly.setdefault(month, {"Audit": 0.0, "Foundation": 0.0, "Ongoing": 0.0})
        if "Audit" in a["activity_type"]:
            monthly[month]["Audit"] += amount
            audit_total += amount
        elif "Foundation" in a["activity_type"]:
            monthly[month]["Foundation"] += amount
            foundation_total += amount
        elif "Ongoing" in a["activity_type"]:
            monthly[month]["Ongoing"] += amount
            ongoing_total += amount
    add_table(ws, r, 0, len(revenue_events), columns, "RevenueEvents", style="Table Style Medium 9")

    r2 = first_data + max(len(revenue_events), 1) + 3
    ws.write(r2, 0, "Monthly totals", fmts["h2"])
    r2 += 1
    ws.write(r2, 0, "Month", fmts["header"])
    ws.write(r2, 1, "Audit", fmts["header"])
    ws.write(r2, 2, "Foundation", fmts["header"])
    ws.write(r2, 3, "Ongoing (new starts)", fmts["header"])
    ws.write(r2, 4, "Total", fmts["header"])
    for i, month in enumerate(sorted(monthly.keys())):
        rr = r2 + 1 + i
        vals = monthly[month]
        ws.write(rr, 0, month, fmts["cell"])
        ws.write_number(rr, 1, vals["Audit"], fmts["cell_money"])
        ws.write_number(rr, 2, vals["Foundation"], fmts["cell_money"])
        ws.write_number(rr, 3, vals["Ongoing"], fmts["cell_money"])
        ws.write_number(rr, 4, sum(vals.values()), fmts["cell_money"])

    r3 = r2 + len(monthly) + 3
    ws.write(r3, 0, "Totals to date", fmts["h2"])
    r3 += 1
    for i, (label, total) in enumerate([("Audit revenue", audit_total), ("Foundation revenue", foundation_total),
                                         ("Ongoing (new-start) revenue", ongoing_total),
                                         ("Total", audit_total + foundation_total + ongoing_total)]):
        ws.write(r3 + i, 0, label, fmts["cell_bold"])
        ws.write_number(r3 + i, 1, total, fmts["cell_money"])

    return ws, audit_total, foundation_total, ongoing_total


def build_import_log_sheet(wb, fmts, tracker):
    ws = wb.add_worksheet("Import Log")
    ws.hide_gridlines(2)
    ws.write(0, 0, "Import Log - conflicts and warnings from every tools/tracker/import_tracker.py run, "
                    "newest first. Source-truth vs manual-truth disagreements are recorded here, never "
                    "silently resolved.", fmts["subtitle"])
    first_row = 2
    columns = [("timestamp", "Timestamp", 22), ("type", "Type", 14), ("message", "Message", 100)]
    log = list(reversed(tracker.get("import_log", [])))

    def classify(message):
        if "older than the stored canonical" in message:
            return "Conflict"
        if "orphaned" in message:
            return "Orphaned"
        if "could not parse" in message:
            return "Parse error"
        return "Note"

    for i, entry in enumerate(log):
        rr = first_row + 1 + i
        ws.write(rr, 0, entry.get("timestamp", ""), fmts["cell"])
        ws.write(rr, 1, classify(entry.get("message", "")), fmts["cell"])
        ws.write(rr, 2, entry.get("message", ""), fmts["cell_wrap"])
    add_table(ws, first_row, 0, len(log), columns, "ImportLog", style="Table Style Medium 9")
    ws.conditional_format(first_row + 1, 1, first_row + max(len(log), 1), 1, {
        "type": "text", "criteria": "containing", "value": "Conflict",
        "format": wb.add_format({"bg_color": REVIEW_BG}),
    })
    return ws


# ---------------------------------------------------------------------------
# Prospects sheet - one canonical row per prospect_id, plus the live
# cadence formulas that turn Activities + Settings into Current Stage /
# Next Action / Next Action Due Date.
# ---------------------------------------------------------------------------

def build_prospects_sheet(wb, fmts, rows, today):
    ws = wb.add_worksheet("Prospects")
    ws.hide_gridlines(2)
    ws.write(0, 0, "Prospects - one row per prospect. Research columns refresh from the latest campaign "
                    "on every rebuild; Do Not Contact (Manual) and Notes are yours to edit; everything from "
                    "Research Pipeline Stage onward is a live formula.", fmts["subtitle"])
    ws.set_row(0, 16)
    first_row = 2
    ws.freeze_panes(first_row + 1, PI["current_stage"])

    for i, row in enumerate(rows):
        rr = first_row + 1 + i
        ws.write(rr, PI["prospect_id"], row["prospect_id"], fmts["cell"])
        ws.write(rr, PI["business"], row["business"], fmts["cell_bold"])
        ws.write(rr, PI["contact_person"], row["contact_person"], fmts["cell"])
        ws.write(rr, PI["contact_email"], row["contact_email"], fmts["cell"])
        # current_stage / next_action / next_action_due_date written after
        # the formula pass below (they depend on the helper columns).
        ws.write(rr, PI["campaign_id"], row["campaign_id"], fmts["cell"])
        ws.write(rr, PI["sector"], row["sector"], fmts["cell"])
        ws.write(rr, PI["geography"], row["geography"], fmts["cell"])
        ws.write(rr, PI["area"], row["area"], fmts["cell_wrap"])
        ws.write(rr, PI["priority"], row["priority"], fmts["cell_center"])
        ws.write(rr, PI["opportunity_type"], row["opportunity_type"], fmts["cell"])
        ws.write(rr, PI["accessibility"], row["accessibility"], fmts["cell"])
        ws.write(rr, PI["role"], row["role"], fmts["cell_wrap"])
        ws.write(rr, PI["contact_phone"], row["contact_phone"], fmts["cell"])
        ws.write(rr, PI["decision_maker_linkedin"], row["decision_maker_linkedin"], fmts["cell"])
        ws.write(rr, PI["website"], row["website"], fmts["cell"])
        ws.write(rr, PI["ready_to_email"], row["ready_to_email"], fmts["cell_center"])
        if row["outreach_rank"] not in (None, ""):
            ws.write_number(rr, PI["outreach_rank"], row["outreach_rank"], fmts["cell_center"])
        else:
            ws.write_blank(rr, PI["outreach_rank"], None, fmts["cell_center"])
        ws.write_boolean(rr, PI["do_not_contact_manual"], row["do_not_contact_manual"], fmts["input_cell"])
        ws.write(rr, PI["notes"], row["notes"], fmts["input_cell"])
        ws.write(rr, PI["research_pipeline_stage"], row["research_pipeline_stage"], fmts["cell"])
        ws.write_boolean(rr, PI["withheld_at_outreach"], row["withheld_at_outreach"], fmts["cell_center"])

    n_data_rows = len(rows)
    last_row, last_col = add_table(ws, first_row, 0, n_data_rows, PROSPECTS_COLUMNS, "Prospects",
                                    style="Table Style Medium 9")

    # --- formula columns ---
    # Plain cell/range references throughout, not Table structured
    # references - see the module-level layout comment above ACTIVITIES_
    # HEADER_ROW for why.
    ACT_PID, ACT_DATE, ACT_TYPE = act_range("prospect_id"), act_range("activity_date"), act_range("activity_type")
    CR_TYPE = cadence_range("label")
    CR_NEXT_ACTION, CR_DAYS = cadence_range("next_action_label"), cadence_range("cadence_days")
    CR_STAGE, CR_BLOCKS = cadence_range("stage_label"), cadence_range("blocks_outreach")

    def C(key, xr):
        return f"{col_letter(PI[key])}{xr}"

    for i, row in enumerate(rows):
        rr = first_row + 1 + i
        xr = rr + 1  # 1-indexed Excel row, for same-row cell references

        f_last_type = (f'=IFERROR(LOOKUP(2,1/(({ACT_PID}={C("prospect_id", xr)})*'
                       f'({ACT_DATE}={C("last_activity_date", xr)})),{ACT_TYPE}),"")')
        ws.write_formula(rr, PI["last_activity_type"], f_last_type, fmts["cell"], row["last_activity_type"])

        # SUMPRODUCT(MAX(...)), not MAXIFS: reproduced and confirmed while
        # building this workbook - MAXIFS over this cross-sheet range
        # intermittently returns blank specifically during Excel's forced
        # recalculation on file open (the cached value and a standalone
        # Evaluate() of the identical MAXIFS formula are both correct; only
        # the on-load recalculation gets it wrong), while this SUMPRODUCT
        # form computes correctly in the same situation. Never revert to
        # MAXIFS here without re-verifying against a real Excel open.
        # Wrapped in an explicit =0 check, not IFERROR: a no-match result
        # is 0 (a valid number, not an error) for both MAXIFS and this
        # SUMPRODUCT form, and 0 as a date serial renders as 1899-12-30,
        # not blank.
        raw_max = f'SUMPRODUCT(MAX(({ACT_PID}={C("prospect_id", xr)})*{ACT_DATE}))'
        f_last_date = f'=IF({raw_max}=0,"",{raw_max})'
        ws.write_formula(rr, PI["last_activity_date"], f_last_date,
                          fmts["cell_date"] if row["last_activity_date"] else fmts["cell"],
                          excel_serial(row["last_activity_date"]) if row["last_activity_date"] else "")

        f_blocked = (f'=OR({C("do_not_contact_manual", xr)}=TRUE,SUMPRODUCT(({ACT_PID}={C("prospect_id", xr)})*'
                     f'ISNUMBER(MATCH({ACT_TYPE},IF({CR_BLOCKS}=TRUE,{CR_TYPE}),0)))>0)')
        ws.write_formula(rr, PI["blocked"], f_blocked, fmts["cell_center"], row["blocked"])

        f_cna = (f'=IF({C("last_activity_type", xr)}="","",IFERROR(INDEX({CR_NEXT_ACTION},'
                 f'MATCH({C("last_activity_type", xr)},{CR_TYPE},0)),""))')
        ws.write_formula(rr, PI["cadence_next_action_raw"], f_cna, fmts["cell"], row["cadence_next_action_raw"])

        f_cd = (f'=IF({C("last_activity_type", xr)}="","",IFERROR(INDEX({CR_DAYS},'
                f'MATCH({C("last_activity_type", xr)},{CR_TYPE},0)),""))')
        ws.write_formula(rr, PI["cadence_days_raw"], f_cd, fmts["cell_center"],
                          row["cadence_days_raw"] if row["cadence_days_raw"] is not None else "")

        f_cs = (f'=IF({C("last_activity_type", xr)}="","",IFERROR(INDEX({CR_STAGE},'
                f'MATCH({C("last_activity_type", xr)},{CR_TYPE},0)),""))')
        ws.write_formula(rr, PI["cadence_stage_raw"], f_cs, fmts["cell"], row["cadence_stage_raw"])

        f_raw_due = (f'=IF(OR({C("last_activity_date", xr)}="",{C("cadence_days_raw", xr)}=""),"",'
                     f'{C("last_activity_date", xr)}+{C("cadence_days_raw", xr)})')
        ws.write_formula(rr, PI["raw_due_date"], f_raw_due,
                          fmts["cell_date"] if row["raw_due_date"] else fmts["cell"],
                          excel_serial(row["raw_due_date"]) if row["raw_due_date"] else "")

        f_due = (f'=IF({C("blocked", xr)}=TRUE,"",'
                 f'IF({C("last_activity_type", xr)}="",'
                 f'IF(AND({C("ready_to_email", xr)}="YES",{C("withheld_at_outreach", xr)}<>TRUE),TODAY(),""),'
                 f'IF({C("raw_due_date", xr)}="","",'
                 f'IF(WEEKDAY({C("raw_due_date", xr)},2)=6,{C("raw_due_date", xr)}+2,'
                 f'IF(WEEKDAY({C("raw_due_date", xr)},2)=7,{C("raw_due_date", xr)}+1,{C("raw_due_date", xr)})))))')
        ws.write_formula(rr, PI["next_action_due_date"], f_due,
                          fmts["cell_date"] if row["next_action_due_date"] else fmts["cell"],
                          excel_serial(row["next_action_due_date"]) if row["next_action_due_date"] else "")

        f_overdue = f'=AND({C("blocked", xr)}<>TRUE,{C("next_action_due_date", xr)}<>"",{C("next_action_due_date", xr)}<TODAY())'
        ws.write_formula(rr, PI["overdue"], f_overdue, fmts["cell_center"], row["overdue"])

        f_due_today = f'=AND({C("blocked", xr)}<>TRUE,{C("next_action_due_date", xr)}<>"",{C("next_action_due_date", xr)}=TODAY())'
        ws.write_formula(rr, PI["due_today"], f_due_today, fmts["cell_center"], row["due_today"])

        f_stage = (f'=IF({C("blocked", xr)}=TRUE,"Do Not Contact",'
                   f'IF({C("last_activity_type", xr)}="",'
                   f'IF(AND({C("ready_to_email", xr)}="YES",{C("withheld_at_outreach", xr)}<>TRUE),"Outreach ready",'
                   f'IF({C("research_pipeline_stage", xr)}="RESEARCHED","Research complete",'
                   f'IF({C("research_pipeline_stage", xr)}="","","Qualified"))),'
                   f'IF(AND(OR({C("cadence_stage_raw", xr)}="Prepared",{C("cadence_stage_raw", xr)}="Contacted"),'
                   f'{C("overdue", xr)}=TRUE),"Follow-up due",{C("cadence_stage_raw", xr)})))')
        ws.write_formula(rr, PI["current_stage"], f_stage, fmts["cell_bold"], row["current_stage"])

        f_next_action = (f'=IF({C("blocked", xr)}=TRUE,"",'
                          f'IF({C("last_activity_type", xr)}="",'
                          f'IF(AND({C("ready_to_email", xr)}="YES",{C("withheld_at_outreach", xr)}<>TRUE),"Send Email 1",""),'
                          f'{C("cadence_next_action_raw", xr)}))')
        ws.write_formula(rr, PI["next_action"], f_next_action, fmts["cell"], row["next_action"])

    ws.data_validation(first_row + 1, PI["do_not_contact_manual"], last_row, PI["do_not_contact_manual"],
                        {"validate": "list", "source": ["TRUE", "FALSE"]})

    # Conditional formatting - whole-row highlight keyed off the boolean/stage helper columns.
    first_data, last_data = first_row + 1, last_row
    rng = f"{col_letter(0)}{first_data + 1}:{col_letter(last_col)}{last_data + 1}"
    stage_col = col_letter(PI["current_stage"])
    overdue_col = col_letter(PI["overdue"])
    due_today_col = col_letter(PI["due_today"])
    blocked_col = col_letter(PI["blocked"])
    priority_col = col_letter(PI["priority"])

    def row_rule(formula, bg):
        ws.conditional_format(rng, {"type": "formula", "criteria": formula,
                                     "format": wb.add_format({"bg_color": bg})})

    row_rule(f"=${blocked_col}{first_data + 1}=TRUE", BLOCKED_BG)
    row_rule(f'=AND(${blocked_col}{first_data + 1}<>TRUE,${overdue_col}{first_data + 1}=TRUE)', OVERDUE_BG)
    row_rule(f'=AND(${blocked_col}{first_data + 1}<>TRUE,${due_today_col}{first_data + 1}=TRUE)', DUE_TODAY_BG)
    row_rule(f'=${stage_col}{first_data + 1}="Replied"', REPLIED_BG)
    row_rule(f'=OR(${stage_col}{first_data + 1}="Audit",${stage_col}{first_data + 1}="Foundation",'
             f'${stage_col}{first_data + 1}="Ongoing")', WON_BG)
    row_rule(f'=${stage_col}{first_data + 1}="Lost"', LOST_BG)
    row_rule(f'=${stage_col}{first_data + 1}="Outreach ready"', READY_BG)
    row_rule(f'=${priority_col}{first_data + 1}="REVIEW"', REVIEW_BG)

    # Collapsible helper-column group (raw cadence lookups + booleans) -
    # visible and auditable, but out of the way day to day.
    for c in range(HELPER_COL_START, last_col + 1):
        width = PROSPECTS_COLUMNS[c][2]
        ws.set_column(c, c, width, None, {"level": 1, "collapsed": True})
    ws.outline_settings(symbols_below=False, symbols_right=True)

    return ws, first_row, last_row


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def build_workbook(tracker, existing_activities, manual_fields, out_path, today=None):
    today = today or date.today()
    cadence_table = cad.default_cadence_table()
    cadence_map = cadence_by_label(cadence_table)

    activities, n_auto_prepared = merge_activities(existing_activities, tracker)
    rows = build_prospect_rows(tracker, activities, manual_fields, today)
    business_by_prospect_id = {row["prospect_id"]: row["business"] for row in rows}

    wb = xlsxwriter.Workbook(out_path, {"default_date_format": "yyyy-mm-dd"})
    fmts = make_formats(wb)

    build_today_sheet(wb, fmts, rows, today)
    build_prospects_sheet(wb, fmts, rows, today)
    wb.define_name("ProspectIDList", "=Prospects!Prospects[Prospect ID]")
    wb.define_name("ActivityTypeList", "=Settings!CadenceRules[Activity Type]")
    build_activities_sheet(wb, fmts, activities, business_by_prospect_id, len(rows))
    build_campaigns_sheet(wb, fmts, tracker, rows)
    build_pipeline_sheet(wb, fmts, rows, tracker)
    build_revenue_sheet(wb, fmts, activities, business_by_prospect_id, cadence_map)
    build_settings_sheet(wb, fmts, cadence_table)
    build_import_log_sheet(wb, fmts, tracker)

    wb.worksheets()[0].activate()
    wb.close()
    return {
        "n_prospects": len(rows),
        "n_activities": len(activities),
        "n_auto_prepared_logged": n_auto_prepared,
        "n_campaigns": len(tracker.get("campaigns", {})),
    }


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--tracker-dir", default=str(Path.home() / "wardith-runs" / "tracker"),
                     help="Where tracker.json/tracker.csv live. Default: ~/wardith-runs/tracker")
    ap.add_argument("--output", default=None,
                     help="Path to write wardith-crm.xlsx. Default: <tracker-dir>/wardith-crm.xlsx")
    ap.add_argument("--today", default=None, help="Override today's date (YYYY-MM-DD), for testing.")
    args = ap.parse_args(argv)

    tracker_dir = Path(args.tracker_dir).expanduser()
    tracker_path = tracker_dir / "tracker.json"
    out_path = Path(args.output).expanduser() if args.output else tracker_dir / "wardith-crm.xlsx"
    today = datetime.strptime(args.today, "%Y-%m-%d").date() if args.today else date.today()

    if not tracker_path.is_file():
        print(f"error: tracker.json not found: {tracker_path}", file=sys.stderr)
        return 1

    tracker = load_tracker(tracker_path)
    existing_activities, manual_fields = load_existing_workbook(out_path)

    tmp_path = out_path.with_suffix(".xlsx.tmp")
    stats = build_workbook(tracker, existing_activities, manual_fields, str(tmp_path), today=today)
    tmp_path.replace(out_path)

    print(f"Wrote {out_path}")
    print(f"Campaigns: {stats['n_campaigns']}")
    print(f"Prospects: {stats['n_prospects']}")
    print(f"Activities: {stats['n_activities']} ({stats['n_auto_prepared_logged']} auto-logged this run)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
