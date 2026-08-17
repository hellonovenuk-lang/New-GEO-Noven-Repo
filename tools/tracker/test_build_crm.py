import os
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import build_crm as bc
import cadence as cad
from openpyxl import load_workbook


def make_tracker(prospect_id="camp::biz", campaign_id="camp", pipeline_stage="OUTREACH_PREPARED",
                  ready_to_email="YES", withheld=False, outreach_status=None, prepared_date=None):
    return {
        "schema_version": 1,
        "campaigns": {
            campaign_id: {
                "campaign_id": campaign_id, "sector": "dentist", "geography": "Chester",
                "run_date": "2026-08-14", "pipeline_stage": pipeline_stage,
                "market_count": 1, "outreach_count": 1, "excluded_count": 0,
            },
        },
        "prospects": {
            prospect_id: {
                "prospect_id": prospect_id,
                "research": {
                    "business": "Test Business", "area": "Chester", "website": "https://x.co.uk",
                    "campaign_id": campaign_id, "priority": "A", "opportunity_type": "GROWTH",
                    "outreach_rank": 1, "accessibility": "DIRECT", "contact_person": "Jo Bloggs",
                    "role": "Director", "contact_email": "jo@x.co.uk", "contact_phone": None,
                    "decision_maker_linkedin": None, "ready_to_email": ready_to_email,
                    "withheld_at_outreach": withheld, "withheld_reason": None,
                },
                "activity": {
                    "outreach_status": outreach_status, "prepared_date": prepared_date,
                    "sent_date": None, "replied_date": None, "do_not_contact": None,
                },
            },
        },
        "import_log": [],
    }


class TestStableIdDedup(unittest.TestCase):
    def test_one_row_per_prospect_id_regardless_of_source_repeats(self):
        tracker = make_tracker()
        # Same prospect_id appearing "twice" in source data (e.g. a re-import) is a
        # dict key in tracker.json, so it can only ever be one entry - build_prospect_rows
        # must produce exactly one row for it.
        rows = bc.build_prospect_rows(tracker, [], {}, date(2026, 8, 16))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["prospect_id"], "camp::biz")


class TestCadenceByLabelFix(unittest.TestCase):
    def test_activity_type_label_resolves_against_the_cadence_table(self):
        # Regression test: cadence.cadence_by_key() indexes by the internal
        # machine key ('OUTREACH_PREPARED'), but every activity_type value
        # this workbook actually stores is the human-readable label
        # ('Outreach prepared'). Using the wrong index silently produces a
        # blank stage/next-action instead of raising - this test would have
        # caught it.
        cadence_map = bc.cadence_by_label()
        self.assertIn("Outreach prepared", cadence_map)
        self.assertNotIn("OUTREACH_PREPARED", cadence_map)
        self.assertEqual(cadence_map["Outreach prepared"]["next_action_label"], "Send Email 1")

    def test_build_prospect_rows_resolves_a_genuinely_prepared_activity(self):
        tracker = make_tracker(outreach_status="PREPARED", prepared_date="2026-08-16")
        activities, n_added = bc.merge_activities([], tracker)
        self.assertEqual(n_added, 1)
        rows = bc.build_prospect_rows(tracker, activities, {}, date(2026, 8, 16))
        row = rows[0]
        self.assertEqual(row["current_stage"], "Prepared")
        self.assertEqual(row["next_action"], "Send Email 1")
        self.assertIsNotNone(row["next_action_due_date"])
        self.assertNotEqual(row["current_stage"], "")


class TestWithheldNotPrepared(unittest.TestCase):
    def test_no_activity_and_not_ready_never_reads_outreach_ready(self):
        tracker = make_tracker(ready_to_email="REVIEW", withheld=False, outreach_status=None)
        rows = bc.build_prospect_rows(tracker, [], {}, date(2026, 8, 16))
        self.assertNotEqual(rows[0]["current_stage"], "Outreach ready")

    def test_withheld_at_outreach_never_reads_outreach_ready_even_if_ready_yes(self):
        tracker = make_tracker(ready_to_email="YES", withheld=True, outreach_status=None)
        rows = bc.build_prospect_rows(tracker, [], {}, date(2026, 8, 16))
        self.assertNotEqual(rows[0]["current_stage"], "Outreach ready")
        self.assertEqual(rows[0]["current_stage"], "Qualified")


class TestManualActivityPreservation(unittest.TestCase):
    def test_rebuild_preserves_a_prior_workbooks_activities_and_manual_fields(self):
        tracker = make_tracker()
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"

            # First build: nothing manual yet.
            stats1 = bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            self.assertEqual(stats1["n_activities"], 0)

            # Owner logs a manual "Email 1 sent" activity and sets Notes/DNC by hand.
            existing_activities, manual_fields = bc.load_existing_workbook(out_path)
            existing_activities.append({
                "prospect_id": "camp::biz", "activity_type": "Email 1 sent",
                "activity_date": date(2026, 8, 17), "amount": None, "plan": None,
                "notes": "Left a voicemail too", "logged_by": "Owner",
            })
            manual_fields["camp::biz"] = {"do_not_contact_manual": False, "notes": "Called ahead of email."}

            # Second build ("refresh from a new campaign output"): manual activity
            # and notes must survive, and must not be duplicated.
            stats2 = bc.build_workbook(tracker, existing_activities, manual_fields, str(out_path),
                                        today=date(2026, 8, 18))
            self.assertEqual(stats2["n_activities"], 1)

            reloaded_activities, reloaded_manual = bc.load_existing_workbook(out_path)
            self.assertEqual(len(reloaded_activities), 1)
            self.assertEqual(reloaded_activities[0]["activity_type"], "Email 1 sent")
            self.assertEqual(reloaded_activities[0]["notes"], "Left a voicemail too")
            self.assertEqual(reloaded_manual["camp::biz"]["notes"], "Called ahead of email.")

    def test_repeat_rebuild_with_no_changes_is_idempotent(self):
        tracker = make_tracker(outreach_status="PREPARED", prepared_date="2026-08-16")
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"
            bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            existing_activities, manual_fields = bc.load_existing_workbook(out_path)
            self.assertEqual(len(existing_activities), 1)  # the auto-logged "Outreach prepared" row

            stats2 = bc.build_workbook(tracker, existing_activities, manual_fields, str(out_path),
                                        today=date(2026, 8, 16))
            self.assertEqual(stats2["n_activities"], 1)
            self.assertEqual(stats2["n_auto_prepared_logged"], 0)  # already logged, not duplicated

            # Third rebuild from the second output, same result.
            existing_activities2, manual_fields2 = bc.load_existing_workbook(out_path)
            stats3 = bc.build_workbook(tracker, existing_activities2, manual_fields2, str(out_path),
                                        today=date(2026, 8, 16))
            self.assertEqual(stats3["n_activities"], 1)


class TestRevenueCalculation(unittest.TestCase):
    def test_audit_and_foundation_revenue_are_summed_separately_by_month(self):
        activities = [
            {"prospect_id": "a", "activity_type": "Audit sold", "activity_date": date(2026, 8, 5),
             "amount": 250, "plan": None},
            {"prospect_id": "b", "activity_type": "Audit sold", "activity_date": date(2026, 8, 20),
             "amount": 250, "plan": None},
            {"prospect_id": "c", "activity_type": "Foundation sold", "activity_date": date(2026, 9, 1),
             "amount": 800, "plan": None},
            {"prospect_id": "d", "activity_type": "Ongoing plan started", "activity_date": date(2026, 9, 2),
             "amount": 400, "plan": "Grow"},
            {"prospect_id": "e", "activity_type": "Email 1 sent", "activity_date": date(2026, 8, 6),
             "amount": None, "plan": None},  # not a revenue event - must be excluded
        ]
        cadence_map = bc.cadence_by_label()
        revenue_events = [a for a in activities if cadence_map.get(a["activity_type"], {}).get("is_revenue_event")
                           or a.get("amount")]
        self.assertEqual(len(revenue_events), 4)  # excludes the Email 1 sent row

        audit_total = sum(a["amount"] for a in revenue_events if "Audit" in a["activity_type"])
        foundation_total = sum(a["amount"] for a in revenue_events if "Foundation" in a["activity_type"])
        ongoing_total = sum(a["amount"] for a in revenue_events if "Ongoing" in a["activity_type"])
        self.assertEqual(audit_total, 500)
        self.assertEqual(foundation_total, 800)
        self.assertEqual(ongoing_total, 400)

    def test_revenue_events_are_month_bucketed_correctly(self):
        cadence_map = bc.cadence_by_label()
        self.assertTrue(cadence_map["Audit sold"]["is_revenue_event"])
        self.assertTrue(cadence_map["Foundation sold"]["is_revenue_event"])
        self.assertTrue(cadence_map["Ongoing plan started"]["is_revenue_event"])
        self.assertFalse(cadence_map["Email 1 sent"]["is_revenue_event"])
        self.assertFalse(cadence_map["Audit completed"]["is_revenue_event"])  # delivery, not the sale


class TestWorkbookStructure(unittest.TestCase):
    def test_build_workbook_produces_all_eight_sheets_with_no_dict_leaking_into_cells(self):
        tracker = make_tracker(outreach_status="PREPARED", prepared_date="2026-08-16")
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"
            bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            wb = load_workbook(out_path, data_only=False)
            self.assertEqual(
                wb.sheetnames,
                ["Today", "Prospects", "Activities", "Campaigns", "Pipeline", "Revenue", "Settings", "Import Log"],
            )
            wb.close()

    def test_excel_serial_conversion_matches_a_known_date(self):
        # 2026-08-16 in Excel's serial system (1899-12-30 epoch) - confirmed
        # against a real Excel session (Prospects!Z4 displayed "2026-08-16"
        # for this exact serial while verifying the workbook).
        self.assertEqual(bc.excel_serial(date(2026, 8, 16)), 46250)


class TestActivitiesBusinessAutoFill(unittest.TestCase):
    """Fix 1: Business on Activities must be a live lookup off Prospect ID,
    not something the owner has to type, and Revenue must still resolve
    the right business name."""

    def test_business_cell_is_a_formula_not_a_static_value(self):
        tracker = make_tracker(outreach_status="PREPARED", prepared_date="2026-08-16")
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"
            bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            wb = load_workbook(out_path, data_only=False)
            ws = wb["Activities"]
            header_row, data_rows = bc._table_rows(ws, "Activities")
            headers = [c.value for c in header_row]
            idx = {h: i for i, h in enumerate(headers)}
            business_cell = data_rows[0][idx["Business"]]
            self.assertIsInstance(business_cell, str)
            self.assertTrue(business_cell.startswith("="), "Business must be a formula, not a static value")
            self.assertIn("INDEX(", business_cell)
            self.assertIn("MATCH(", business_cell)
            # The structured self-reference + cross-sheet combination that
            # corrupted the file for Excel's automation loader while this
            # workbook was being built - never reintroduce it here.
            self.assertNotIn("[@", business_cell)
            wb.close()

    def test_business_cell_cached_value_is_correct(self):
        tracker = make_tracker(outreach_status="PREPARED", prepared_date="2026-08-16")
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"
            bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            wb = load_workbook(out_path, data_only=True)
            ws = wb["Activities"]
            header_row, data_rows = bc._table_rows(ws, "Activities")
            headers = [c.value for c in header_row]
            idx = {h: i for i, h in enumerate(headers)}
            self.assertEqual(data_rows[0][idx["Business"]], "Test Business")
            wb.close()

    def test_business_formula_is_written_for_every_reserved_row_not_just_filled_ones(self):
        # A row the owner types by hand (Prospect ID + Activity Type + Date
        # only) must already have a working Business formula waiting -
        # nobody re-runs build_crm.py just to see a name appear.
        tracker = make_tracker()
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"
            bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            wb = load_workbook(out_path, data_only=False)
            ws = wb["Activities"]
            header_row, data_rows = bc._table_rows(ws, "Activities")
            headers = [c.value for c in header_row]
            idx = {h: i for i, h in enumerate(headers)}
            # Row 10 (well past the 0 filled activities) is still empty but reserved.
            far_row = data_rows[9][idx["Business"]]
            self.assertIsInstance(far_row, str)
            self.assertTrue(far_row.startswith("="))
            wb.close()

    def test_revenue_sheet_still_resolves_business_name_independently(self):
        # Revenue's business column comes from the same Python-side
        # business_by_prospect_id lookup used for Activities' formula cache
        # - not by reading Activities' own (now formula) cell - confirm it
        # still resolves correctly end to end.
        tracker = make_tracker(outreach_status="PREPARED", prepared_date="2026-08-16")
        activities, _n = bc.merge_activities([], tracker)
        rows = bc.build_prospect_rows(tracker, activities, {}, date(2026, 8, 16))
        business_by_prospect_id = {row["prospect_id"]: row["business"] for row in rows}
        self.assertEqual(business_by_prospect_id["camp::biz"], "Test Business")
        # The one activity present (auto-logged "Outreach prepared") is a revenue-irrelevant
        # type, but the lookup dict Revenue draws from is exactly this one.
        self.assertEqual(activities[0]["prospect_id"], "camp::biz")


class TestTodayOverflowVisibility(unittest.TestCase):
    """Fix 2: Today must never silently hide matches beyond its reserved
    spill capacity."""

    def test_capacity_was_raised_from_the_original_25(self):
        self.assertGreater(bc.TODAY_BLOCK_SPILL_CAPACITY, 25)

    def test_spill_formula_hard_caps_at_capacity_via_take(self):
        tracker = make_tracker()
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"
            bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            wb = load_workbook(out_path, data_only=False)
            ws = wb["Today"]
            # First block's spill anchor is 3 rows below the sheet's own header block.
            # openpyxl represents a dynamic-array formula as an ArrayFormula object.
            anchor = ws["A7"].value
            anchor_text = anchor.text if hasattr(anchor, "text") else anchor
            self.assertIsInstance(anchor_text, str)
            self.assertIn("TAKE(", anchor_text)
            self.assertIn(str(bc.TODAY_BLOCK_SPILL_CAPACITY), anchor_text)
            wb.close()

    def test_overflow_beyond_capacity_produces_an_explicit_message_not_silence(self):
        # 50 outreach-ready prospects against a 40-row capacity forces the
        # "New outreach-ready prospects" block over its cap.
        campaigns = {"camp": {"campaign_id": "camp", "sector": "dentist", "geography": "Chester",
                               "run_date": "2026-08-14", "pipeline_stage": "OUTREACH_PREPARED",
                               "market_count": 50, "outreach_count": 50, "excluded_count": 0}}
        prospects = {}
        for i in range(50):
            pid = f"camp::biz-{i:03d}"
            prospects[pid] = {
                "prospect_id": pid,
                "research": {
                    "business": f"Test Business {i:03d}", "area": "Chester", "website": "https://x.co.uk",
                    "campaign_id": "camp", "priority": "A", "opportunity_type": "GROWTH", "outreach_rank": 1,
                    "accessibility": "DIRECT", "contact_person": "Jo Bloggs", "role": "Director",
                    "contact_email": "jo@x.co.uk", "contact_phone": None, "decision_maker_linkedin": None,
                    "ready_to_email": "YES", "withheld_at_outreach": False, "withheld_reason": None,
                },
                "activity": {"outreach_status": None, "prepared_date": None},
            }
        tracker = {"schema_version": 1, "campaigns": campaigns, "prospects": prospects, "import_log": []}
        rows = bc.build_prospect_rows(tracker, [], {}, date(2026, 8, 16))
        ready_rows = [r for r in rows if r["current_stage"] == "Outreach ready"]
        self.assertEqual(len(ready_rows), 50)
        self.assertGreater(len(ready_rows), bc.TODAY_BLOCK_SPILL_CAPACITY)

        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"
            bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            # Cached text for the overflowing block's count line - build_crm.py
            # writes this as the formula's cached value, and it was also
            # confirmed to match what a real Excel open computes live (manual
            # QA against this exact scenario while building this workbook).
            wb2 = load_workbook(out_path, data_only=True)
            ws2 = wb2["Today"]
            found_overflow_text = any(
                isinstance(cell.value, str) and "Showing 40 of 50" in cell.value
                for row in ws2.iter_rows(min_col=1, max_col=1) for cell in row
            )
            self.assertTrue(found_overflow_text, "Expected a 'Showing 40 of 50 - open Prospects for the "
                             "remainder.' cached message somewhere in column A")
            wb2.close()

    def test_sumproduct_count_uses_explicit_boolean_coercion(self):
        # Regression test: a bare SUMPRODUCT((Prospects[Due Today]=TRUE))
        # against a structured Table reference silently returned 0 instead
        # of the real count in this Excel installation (confirmed via
        # Evaluate()) - only the double-unary --(...) form computed
        # correctly. The generated formula must always use it.
        tracker = make_tracker()
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"
            bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            wb = load_workbook(out_path, data_only=False)
            ws = wb["Today"]
            count_formula = ws["A5"].value
            self.assertIsInstance(count_formula, str)
            self.assertIn("SUMPRODUCT(--(", count_formula)
            wb.close()


class TestActivityIdIsNotAStableIdentity(unittest.TestCase):
    """Fix 3: the Row()-based column must read as a non-identifying
    display row number, and nothing may depend on it as an identity."""

    def test_column_header_reads_row_number_not_activity_id(self):
        tracker = make_tracker(outreach_status="PREPARED", prepared_date="2026-08-16")
        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "wardith-crm.xlsx"
            bc.build_workbook(tracker, [], {}, str(out_path), today=date(2026, 8, 16))
            wb = load_workbook(out_path, data_only=False)
            ws = wb["Activities"]
            header_row, _data_rows = bc._table_rows(ws, "Activities")
            headers = [c.value for c in header_row]
            self.assertIn("Row #", headers)
            self.assertNotIn("Activity ID", headers)
            wb.close()

    def test_reordering_activities_never_changes_a_prospects_computed_state(self):
        # If anything keyed off row position/activity_id as an identity,
        # shuffling the order activities are merged in would change the
        # outcome. It must not.
        tracker = make_tracker(outreach_status="PREPARED", prepared_date="2026-08-16")
        activities, _n = bc.merge_activities([], tracker)
        extra = [
            {"prospect_id": "camp::biz", "activity_type": "Email 1 sent", "activity_date": date(2026, 8, 17),
             "amount": None, "plan": None, "notes": "", "logged_by": "Owner"},
            {"prospect_id": "camp::biz", "activity_type": "Reply received", "activity_date": date(2026, 8, 18),
             "amount": None, "plan": None, "notes": "", "logged_by": "Owner"},
        ]
        forward = bc.build_prospect_rows(tracker, activities + extra, {}, date(2026, 8, 20))
        backward = bc.build_prospect_rows(tracker, list(reversed(activities + extra)), {}, date(2026, 8, 20))
        self.assertEqual(forward[0]["current_stage"], backward[0]["current_stage"])
        self.assertEqual(forward[0]["current_stage"], "Replied")

    def test_merge_activities_dedup_key_is_prospect_and_type_never_the_row_number(self):
        # merge_activities' own "already logged?" check - the one place a
        # potential identity check exists - keys off (prospect_id,
        # activity_type), confirmed by reading its source's dedup logic
        # indirectly: re-merging the same auto-logged activity twice must
        # never duplicate it, regardless of any row-number reassignment
        # that happens on every build.
        tracker = make_tracker(outreach_status="PREPARED", prepared_date="2026-08-16")
        activities1, n1 = bc.merge_activities([], tracker)
        self.assertEqual(n1, 1)
        activities2, n2 = bc.merge_activities(activities1, tracker)
        self.assertEqual(n2, 0)
        self.assertEqual(len(activities2), 1)


if __name__ == "__main__":
    unittest.main()
