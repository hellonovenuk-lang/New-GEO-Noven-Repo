#!/usr/bin/env python3
"""
Regression tests for build_workbook.py's validate() and rendering logic.
Stdlib unittest + openpyxl (used read-only, to confirm data-only reopening
returns populated calculated values - see README.md).

Run: python3 test_build_workbook.py -v
"""
import copy
import json
import os
import tempfile
import unittest

from openpyxl import load_workbook

import build_workbook as bw
import scoring_engine as se

SAMPLE_PATH = os.path.join(os.path.dirname(__file__), "sample", "sample-campaign.json")


def load_sample():
    with open(SAMPLE_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def scored_fixture():
    """A small, schema-valid, fully-scored campaign (2 outreach businesses,
    one ready, one not) - enough to exercise every sheet builder without
    depending on any real campaign's data."""
    data = {
        "run": {
            "sector": "fictitious trade", "geography": "Sampleford", "campaign_slug": "wb-test", "date": "2026-08-16",
            "questions": [
                {"question_id": "q01", "text": "best A and B providers"},
                {"question_id": "q02", "text": "B only question"},
            ],
            "providers": [{"provider": "openai", "model": "x"}],
            "responses_per_question": 15,
            "service_scopes": [{"label": "combined", "applicable_services": ["A", "B"]}],
            "question_relevance": [
                {"question_id": "q01", "type": "SINGLE_SERVICE_INCLUSIVE", "rationale": "inclusive"},
                {"question_id": "q02", "type": "SERVICE_ONLY", "service": "B", "rationale": "b only"},
            ],
        },
        "market": [],
        "outreach": [
            {
                "priority": "REVIEW", "business": "Ready Co", "area": "Sampleford", "total_ai_appearances": 0,
                "strongest_competitor": "x", "competitor_appearances": 0, "competitive_gap_finding": "x",
                "why_prospect": "x", "legal_entity": "x", "company_number": "1", "company_status": "Active",
                "ready_to_email": "REVIEW", "evidence_source_ids": ["S001"], "accessibility": "DIRECT",
                "service_scope": "combined", "question_appearances": {"q01": 5, "q02": 5},
                "commercial_fit": 5, "service_relevance": 5, "business_credibility": 5, "ability_to_buy": 5,
                "decision_maker_identified": 5, "direct_dm_route": 5, "contact_route_quality": 5,
                "contact_identity_confidence": 5, "research_completeness": 5,
            },
            {
                "priority": "REVIEW", "business": "Not Ready Co", "area": "Sampleford", "total_ai_appearances": 0,
                "strongest_competitor": "x", "competitor_appearances": 0, "competitive_gap_finding": "x",
                "why_prospect": "x", "legal_entity": "x", "company_number": "2", "company_status": "Active",
                "ready_to_email": "REVIEW", "evidence_source_ids": ["S001"], "accessibility": "REVIEW",
                "service_scope": "combined", "question_appearances": {"q01": 1, "q02": 0},
                "commercial_fit": 2, "service_relevance": 2, "business_credibility": 1, "ability_to_buy": 1,
                "decision_maker_identified": 0, "direct_dm_route": 0, "contact_route_quality": 0,
                "contact_identity_confidence": 0, "research_completeness": 1,
            },
        ],
        "excluded": [],
        "sources": [{"source_id": "S001", "business": "x", "publisher": "x", "fact_supported": "x",
                     "url": "x", "access_date": "2026-08-16", "fact_category": "AI_APPEARANCE"}],
    }
    se.run_engine(data)
    return data


class ValidateLegacyFieldsTests(unittest.TestCase):
    """The pre-v2 required fields/enums /outreach depends on - unchanged."""

    def test_sample_campaign_validates_as_shipped(self):
        data = load_sample()
        bw.validate(data)  # must not raise

    def test_outreach_entry_missing_accessibility_fails(self):
        data = load_sample()
        del data["outreach"][0]["accessibility"]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate(data)
        self.assertIn("accessibility", str(ctx.exception))

    def test_outreach_entry_bad_accessibility_value_fails(self):
        data = load_sample()
        data["outreach"][0]["accessibility"] = "SOMEWHERE_ELSE"
        with self.assertRaises(bw.ValidationError):
            bw.validate(data)

    def test_market_entry_accessibility_is_optional(self):
        data = load_sample()
        self.assertNotIn("accessibility", data["market"][0])
        bw.validate(data)  # must not raise

    def test_opportunity_type_review_is_now_valid(self):
        # Section 3a supersedes the old "REVIEW is not an opportunity type"
        # rule for a scored business - confirm the schema-level enum allows it.
        data = load_sample()
        data["outreach"][0]["opportunity_type"] = "REVIEW"
        bw.validate(data)  # must not raise


class ValidateScoringFieldsTests(unittest.TestCase):
    def test_scored_fixture_validates(self):
        data = scored_fixture()
        bw.validate(data)  # must not raise

    def test_service_scope_without_value_fields_fails(self):
        data = scored_fixture()
        del data["outreach"][0]["commercial_fit"]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate(data)
        self.assertIn("commercial_fit", str(ctx.exception))

    def test_service_scope_without_derived_fields_fails(self):
        # Simulates forgetting to run scoring_engine.py before build_workbook.py.
        data = scored_fixture()
        del data["outreach"][0]["final_score"]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate(data)
        self.assertIn("final_score", str(ctx.exception))

    def test_service_scope_requires_run_level_definitions(self):
        data = scored_fixture()
        del data["run"]["service_scopes"]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate(data)
        self.assertIn("service_scopes", str(ctx.exception))

    def test_outreach_rank_on_non_ready_entry_fails(self):
        data = scored_fixture()
        data["outreach"][1]["ready_to_email"] = "REVIEW"
        data["outreach"][1]["outreach_rank"] = 1
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate(data)
        self.assertIn("outreach_rank", str(ctx.exception))

    def test_non_consecutive_outreach_ranks_fail(self):
        data = scored_fixture()
        data["outreach"][0]["ready_to_email"] = "YES"
        data["outreach"][0]["outreach_rank"] = 2  # should be 1, the only ready entry
        with self.assertRaises(bw.ValidationError):
            bw.validate(data)


class RenderTests(unittest.TestCase):
    def test_workbook_builds_all_five_sheets(self):
        data = scored_fixture()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            n = bw.build_workbook(data, out)
            self.assertEqual(n, 1)  # Ready Co only
            self.assertTrue(os.path.exists(out))
            wb = load_workbook(out)
            self.assertEqual(set(wb.sheetnames), {"Methodology", "Scoring", "Shortlist", "Evidence", "QC"})

    def test_sample_campaign_still_renders_without_scoring(self):
        # A campaign with nothing scored (no service_scope anywhere) must
        # still render cleanly - Scoring/Shortlist are simply empty.
        data = load_sample()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.validate(data)
            n = bw.build_workbook(data, out)
            self.assertEqual(n, 0)
            self.assertTrue(os.path.exists(out))

    def test_data_only_reopen_returns_populated_values(self):
        data = scored_fixture()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            wb = load_workbook(out, data_only=True)
            ws = wb["Scoring"]
            headers = [c.value for c in ws[1]]
            idx = {h: i for i, h in enumerate(headers)}
            check_cols = ["Overall rank", "Final qualification score (0-100)",
                          "Relevance-normalized visibility %", "Opportunity type", "Business verified"]
            blanks = 0
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if row[idx["Business"]] is None:
                    continue
                for col in check_cols:
                    if row[idx[col]] is None:
                        blanks += 1
            self.assertEqual(blanks, 0)

    def test_shortlist_contains_only_ready_business(self):
        data = scored_fixture()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            wb = load_workbook(out)
            ws = wb["Shortlist"]
            businesses = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
            # column 0 is "Overall rank" not business - re-fetch by header
            headers = [c.value for c in ws[1]]
            biz_col = headers.index("Business")
            names = [row[biz_col].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row) if row[biz_col].value]
            self.assertEqual(names, ["Ready Co"])

    def test_end_to_end_write_to_real_file(self):
        data = load_sample()
        bw.validate(data)
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            self.assertTrue(os.path.exists(out))
            self.assertGreater(os.path.getsize(out), 0)


def full_cohort_fixture():
    """scored_fixture() plus the run.scoring_cohort / cohort_inclusion_min_appearances
    fields --require-scored expects. Both scored_fixture() businesses meet the
    floor (min_appearances=1) and are both marked SCORED, matching the fact
    that scored_fixture() already gives both of them service_scope."""
    data = scored_fixture()
    data["run"]["cohort_inclusion_min_appearances"] = 1
    data["run"]["scoring_cohort"] = [
        {"business": "Ready Co", "status": "SCORED"},
        {"business": "Not Ready Co", "status": "SCORED"},
    ]
    return data


class LegacyWorkbookLabelingTests(unittest.TestCase):
    """A campaign with nothing scored must render, but be unambiguously
    labelled as legacy/unscored rather than silently rendering an empty
    Scoring/QC block that could be misread as a canonical PASS."""

    def test_methodology_sheet_carries_legacy_banner(self):
        data = load_sample()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            wb = load_workbook(out)
            ws = wb["Methodology"]
            col_a_text = " ".join(c.value for c in ws["A"] if c.value)
            self.assertIn("LEGACY / UNSCORED", col_a_text)

    def test_scored_campaign_has_no_legacy_banner(self):
        data = scored_fixture()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            wb = load_workbook(out)
            ws = wb["Methodology"]
            col_a_text = " ".join(c.value for c in ws["A"] if c.value)
            self.assertNotIn("LEGACY / UNSCORED", col_a_text)

    def test_empty_pool_qc_rows_are_not_applicable_not_pass(self):
        data = load_sample()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            wb = load_workbook(out)
            ws = wb["QC"]
            results = [row[1].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row) if row[1].value]
            self.assertIn("NOT APPLICABLE", results)
            self.assertNotIn("PASS", results)  # no vacuous PASS over the empty pool

    def test_qc_does_not_claim_canonical_pass_for_legacy_campaign(self):
        data = load_sample()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            wb = load_workbook(out)
            ws = wb["QC"]
            all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
            self.assertIn("LEGACY / UNSCORED", all_text)

    def test_scored_campaign_qc_still_reports_pass(self):
        # Confirm the branch split didn't disturb the real scored path.
        data = scored_fixture()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            wb = load_workbook(out)
            ws = wb["QC"]
            results = [row[1].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row) if row[1].value]
            self.assertIn("PASS", results)
            self.assertNotIn("NOT APPLICABLE", results)


class RequireScoredTests(unittest.TestCase):
    """--require-scored: the forward-only gate for a newly-generated
    /qualify campaign. Default (no flag) validate() must stay untouched -
    covered by every test above this class."""

    def test_legacy_campaign_rejected(self):
        data = load_sample()
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        msg = str(ctx.exception)
        self.assertIn("no business", msg)
        self.assertIn("service_scope", msg)

    def test_missing_service_scopes_rejected(self):
        data = full_cohort_fixture()
        del data["run"]["service_scopes"]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("service_scopes", str(ctx.exception))

    def test_missing_question_relevance_rejected(self):
        data = full_cohort_fixture()
        data["run"]["question_relevance"] = data["run"]["question_relevance"][:1]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("question_relevance", str(ctx.exception))

    def test_missing_scoring_cohort_rejected(self):
        data = full_cohort_fixture()
        del data["run"]["scoring_cohort"]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("scoring_cohort", str(ctx.exception))

    def test_missing_cohort_inclusion_floor_rejected(self):
        data = full_cohort_fixture()
        del data["run"]["cohort_inclusion_min_appearances"]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("cohort_inclusion_min_appearances", str(ctx.exception))

    def test_cohort_member_meeting_floor_but_absent_is_rejected(self):
        # Reproduces the "Open Plan Building" failure mode: a business with
        # enough appearances to meet the campaign's own stated floor, but
        # left out of the cohort list entirely, must be named and rejected.
        data = full_cohort_fixture()
        data["market"].append({
            "business": "Open Plan Analogue", "area": "Sampleford", "disposition": "REVIEW",
            "total_ai_appearances": 32,
        })
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("Open Plan Analogue", str(ctx.exception))

    def test_business_below_floor_need_not_be_in_cohort(self):
        data = full_cohort_fixture()
        data["market"].append({
            "business": "Tiny Co", "area": "Sampleford", "disposition": "REVIEW",
            "total_ai_appearances": 0,
        })
        bw.validate_strict_scored(data)  # must not raise - below the floor of 1

    def test_cohort_scored_status_without_service_scope_rejected(self):
        data = full_cohort_fixture()
        data["run"]["scoring_cohort"].append({"business": "Ghost Co", "status": "SCORED"})
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("Ghost Co", str(ctx.exception))

    def test_cohort_excluded_without_reason_rejected_by_validate(self):
        # Shape-checking lives in validate() itself, unconditionally.
        data = full_cohort_fixture()
        data["run"]["scoring_cohort"].append({"business": "Excluded Co", "status": "EXCLUDED"})
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate(data)
        self.assertIn("EXCLUDED", str(ctx.exception))

    def test_cohort_incomplete_without_missing_evidence_rejected_by_validate(self):
        data = full_cohort_fixture()
        data["run"]["scoring_cohort"].append({"business": "Pending Co", "status": "INCOMPLETE"})
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate(data)
        self.assertIn("INCOMPLETE", str(ctx.exception))

    def test_outreach_entry_without_service_scope_rejected(self):
        # No qualitative-only fallback allowed in outreach[] for a new campaign.
        data = full_cohort_fixture()
        data["run"]["scoring_cohort"].append({"business": "Qualitative Co", "status": "SCORED"})
        data["outreach"].append({
            "priority": "REVIEW", "business": "Qualitative Co", "area": "Sampleford",
            "total_ai_appearances": 5, "strongest_competitor": "x", "competitor_appearances": 1,
            "competitive_gap_finding": "x", "why_prospect": "x", "legal_entity": "x",
            "company_number": "9", "company_status": "Active", "ready_to_email": "REVIEW",
            "evidence_source_ids": ["S001"], "accessibility": "DIRECT",
        })
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("Qualitative Co", str(ctx.exception))
        self.assertIn("service_scope", str(ctx.exception))

    def test_ready_without_outreach_rank_rejected(self):
        data = full_cohort_fixture()
        # scored_fixture()'s "Ready Co" already earns ready_to_email=YES from
        # the engine's own gate (perfect VALUE fields) - confirm that, then
        # simulate a hand-edited JSON where outreach_rank never got
        # (re-)computed, e.g. after scoring_engine.py --in-place was skipped
        # on a later edit.
        self.assertEqual(data["outreach"][0]["ready_to_email"], "YES")
        del data["outreach"][0]["outreach_rank"]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("outreach_rank", str(ctx.exception))

    def test_fully_scored_campaign_with_cohort_is_accepted(self):
        data = full_cohort_fixture()
        bw.validate(data)
        bw.validate_strict_scored(data)  # must not raise

    def test_main_cli_rejects_legacy_campaign_and_writes_nothing(self):
        with tempfile.TemporaryDirectory() as tmp:
            in_path = os.path.join(tmp, "campaign.json")
            out_path = os.path.join(tmp, "out.xlsx")
            with open(in_path, "w", encoding="utf-8") as f:
                json.dump(load_sample(), f)
            import subprocess
            import sys as _sys
            result = subprocess.run(
                [_sys.executable, os.path.join(os.path.dirname(__file__), "build_workbook.py"),
                 "--input", in_path, "--output", out_path, "--require-scored"],
                capture_output=True, text=True,
            )
            self.assertNotEqual(result.returncode, 0)
            self.assertFalse(os.path.exists(out_path))
            self.assertIn("canonical scoring requirements not met", result.stderr)


class NarrativeContradictionDetectionTests(unittest.TestCase):
    """detect_narrative_contradictions(): the mechanical checks behind both
    the QC-sheet warning (default render) and the --require-scored reject
    (a new campaign)."""

    def test_clean_generated_entry_has_no_contradictions(self):
        # Dogfooding: the generator's own output must never trip its own
        # detector.
        data = full_cohort_fixture()
        for o in data["outreach"]:
            self.assertEqual(bw.detect_narrative_contradictions(o), [])

    def test_wrong_denominator_in_primary_claim_detected(self):
        data = full_cohort_fixture()
        entry = data["outreach"][0]
        entry["competitive_gap_finding"] = (
            f"{entry['business']} was named in 13 of the 90 raw answers collected across all six questions."
        )
        problems = bw.detect_narrative_contradictions(entry)
        self.assertTrue(any("primary visibility claim" in p for p in problems))

    def test_why_prospect_equal_to_business_type_notes_detected(self):
        data = full_cohort_fixture()
        entry = data["outreach"][0]
        entry["business_type_notes"] = "Kitchen specialist; independently owned local business"
        entry["why_prospect"] = "Kitchen specialist; independently owned local business"
        problems = bw.detect_narrative_contradictions(entry)
        self.assertTrue(any("business_type_notes" in p for p in problems))

    def test_competitor_and_coverage_figures_are_not_falsely_flagged(self):
        # Requirement 4: a per-question or competitor figure using a
        # different, clearly documented denominator must not be confused
        # with the primary visibility claim.
        data = full_cohort_fixture()
        entry = data["outreach"][0]
        entry["competitive_gap_finding"] = (
            f"{entry['business']} appears in {se._fmt_num(entry['relevant_appearances'])} of "
            f"{se._fmt_num(entry['relevant_opportunities'])} relevant opportunities. A competitor sits "
            f"at 40 of 90 raw mentions, and coverage was 5 of 6 questions."
        )
        self.assertEqual(bw.detect_narrative_contradictions(entry), [])

    def test_real_ajk_style_defect_is_detected(self):
        # The actual real-run shape that prompted this whole change,
        # reproduced with fictitious numbers: relevant_opportunities=60,
        # narrative claims "13 of the 90 raw answers".
        entry = {
            "business": "Fixture AJK-style Co", "relevant_appearances": 13, "relevant_opportunities": 60,
            "competitive_gap_finding": "Fixture AJK-style Co was named in 13 of the 90 raw answers collected "
                                        "across all six questions and three assistants.",
            "why_prospect": "Kitchen specialist (design/supply/install); independently owned local business",
            "business_type_notes": "Kitchen specialist (design/supply/install); independently owned local business",
        }
        problems = bw.detect_narrative_contradictions(entry)
        self.assertEqual(len(problems), 2)  # both the denominator mismatch and the business_type_notes copy


class NarrativeQcAndRequireScoredTests(unittest.TestCase):
    """QC-sheet warnings (default render, historical files stay readable)
    vs. --require-scored hard rejection (a new campaign)."""

    def test_qc_flags_never_regenerated_narrative_without_blocking_render(self):
        data = full_cohort_fixture()
        # Simulate a legacy narrative that was hand-typed and never run
        # through scoring_engine.py's mandatory generation step.
        data["outreach"][0]["narrative_generated_from"] = "stale-signature-from-before-a-rescore"
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.validate(data)  # must not raise - default mode stays permissive
            bw.build_workbook(data, out)  # must not raise, workbook still renders
            wb = load_workbook(out)
            ws = wb["QC"]
            rows = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row) if row[0].value}
            self.assertEqual(rows["Narrative fields regenerated from current scored values"], "FAIL")

    def test_qc_reports_pass_for_clean_narratives(self):
        data = full_cohort_fixture()
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            wb = load_workbook(out)
            ws = wb["QC"]
            rows = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row) if row[0].value}
            self.assertEqual(rows["Narrative fields regenerated from current scored values"], "PASS")
            self.assertEqual(rows["Narrative fields consistent with structured data"], "PASS")

    def test_require_scored_rejects_never_regenerated_narrative(self):
        data = full_cohort_fixture()
        data["outreach"][0]["narrative_generated_from"] = "stale-signature-from-before-a-rescore"
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("not (re)generated", str(ctx.exception))

    def test_require_scored_rejects_missing_narrative_generated_from(self):
        data = full_cohort_fixture()
        del data["outreach"][0]["narrative_generated_from"]
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("not (re)generated", str(ctx.exception))

    def test_require_scored_rejects_why_prospect_equal_to_business_type_notes(self):
        data = full_cohort_fixture()
        data["outreach"][0]["business_type_notes"] = "Kitchen specialist; independently owned"
        data["outreach"][0]["why_prospect"] = "Kitchen specialist; independently owned"
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("business_type_notes", str(ctx.exception))

    def test_require_scored_rejects_contradictory_primary_visibility_figure(self):
        data = full_cohort_fixture()
        entry = data["outreach"][0]
        entry["competitive_gap_finding"] = f"{entry['business']} was named in 999 of the 90 raw answers."
        # Re-stamp the signature so this fails on the CONTRADICTION check
        # specifically, not merely the staleness check - isolates the two.
        entry["narrative_generated_from"] = se.narrative_signature(entry)
        with self.assertRaises(bw.ValidationError) as ctx:
            bw.validate_strict_scored(data)
        self.assertIn("primary visibility claim", str(ctx.exception))

    def test_require_scored_accepts_clean_regenerated_narrative(self):
        data = full_cohort_fixture()
        bw.validate_strict_scored(data)  # must not raise - se.run_engine() already generated+stamped it


class ShortlistNarrativeMappingTests(unittest.TestCase):
    """Requirement 5: the Shortlist sheet's "Why this is a prospect" column
    must show why_prospect, never business_type_notes."""

    def test_shortlist_shows_why_prospect_not_business_type_notes(self):
        data = full_cohort_fixture()
        entry = data["outreach"][0]  # "Ready Co" - already ready_to_email=YES from the engine's own gate
        self.assertEqual(entry["ready_to_email"], "YES")
        entry["business_type_notes"] = "This text must NOT appear in the Shortlist prospect-reason column."
        # entry["why_prospect"] is already the engine-generated narrative from full_cohort_fixture()'s
        # se.run_engine() call, distinct from business_type_notes (not part of the narrative signature,
        # so setting it here doesn't trigger regeneration - that's fine, we're checking the render, not
        # generation).
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out.xlsx")
            bw.build_workbook(data, out)
            wb = load_workbook(out)
            ws = wb["Shortlist"]
            headers = [c.value for c in ws[1]]
            why_col = headers.index("Why this is a prospect")
            biz_col = headers.index("Business")
            row = next(r for r in ws.iter_rows(min_row=2, max_row=ws.max_row) if r[biz_col].value == entry["business"])
            self.assertEqual(row[why_col].value, entry["why_prospect"])
            self.assertNotIn("must NOT appear", row[why_col].value)


if __name__ == "__main__":
    unittest.main()
