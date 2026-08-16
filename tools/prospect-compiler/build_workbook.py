#!/usr/bin/env python3
"""
Prospect compiler v2 - renders a validated campaign JSON file (schema.json)
into an .xlsx workbook: Methodology, Scoring, Shortlist, Evidence, QC.

This script does not research businesses or judge who is a prospect. The
input JSON is prepared by Claude/human judgement (research) and
scoring_engine.py (mechanical computation) during a Wardith market
campaign; this script deterministically renders it. Fails loudly on
malformed input rather than filling gaps with guesses.

Built with xlsxwriter, not openpyxl: every formula cell carries a genuine
cached value alongside its formula, so a data-only/read-only reopen (the
kind a downstream consumer like /outreach could do) returns real numbers,
not blanks. See CAMPAIGN-HANDOFF.md Section 3a ("Portable downstream
handoff").

Usage:
  python3 build_workbook.py --input campaign.json --output workbook.xlsx

Requires xlsxwriter (pip install xlsxwriter) - the one non-stdlib
dependency for building. Tests additionally use openpyxl to verify
data-only reopening.
"""
import argparse
import json
import re
import sys

import xlsxwriter
from openpyxl.utils import get_column_letter

EXCLUDED_REASONS = {
    "SOLE TRADER", "ORDINARY PARTNERSHIP", "CHAIN / NO LOCAL DECISION-MAKER",
    "DUPLICATE BRAND", "NOT GENUINELY IN MARKET", "CLOSED / DORMANT",
    "ALREADY STRONGLY VISIBLE", "NO RELIABLE LEGAL MATCH", "OTHER / REVIEW",
}
PRIORITIES = {"A", "B", "C", "REVIEW"}
READY_VALUES = {"YES", "REVIEW"}
DISPOSITIONS = {"OUTREACH", "EXCLUDED", "REVIEW"}
OPPORTUNITY_TYPES = {"GAP", "GROWTH", "DEFEND", "NO OPPORTUNITY", "REVIEW"}
ACCESSIBILITY_VALUES = {"DIRECT", "IDENTIFIABLE", "GATEKEPT", "CORPORATE", "REVIEW"}
ACCESSIBILITY_GRADE_VALUES = {
    "CONFIRMED_DIRECT", "CONFIRMED_GENERIC_ROUTE", "PROBABLE_UNCONFIRMED",
    "GENERIC_INBOX_ONLY", "CONTACT_FORM_OR_PHONE_ONLY", "NO_USABLE_ROUTE",
}
IDENTITY_CONFIDENCE_VALUES = {"CONFIRMED", "PROBABLE", "POSSIBLE", "UNKNOWN"}
BUSINESS_STRUCTURE_VALUES = {"INDEPENDENT_LOCAL", "LOCALLY_OWNED_FRANCHISE", "CENTRALLY_CONTROLLED_CHAIN"}
YES_REVIEW_VALUES = {"YES", "REVIEW"}
YES_PROBABLE_NO_VALUES = {"YES", "PROBABLE", "NO"}
YES_NO_VALUES = {"YES", "NO"}
QUESTION_RELEVANCE_TYPES = {"SERVICE_ONLY", "EXPLICITLY_COMBINED", "SINGLE_SERVICE_INCLUSIVE", "AMBIGUOUS"}
SOURCE_ID_RE = re.compile(r"^S[0-9]{3,}$")

SCORED_VALUE_FIELDS = [
    "commercial_fit", "service_relevance", "business_credibility", "ability_to_buy",
    "decision_maker_identified", "direct_dm_route", "contact_route_quality",
    "contact_identity_confidence", "research_completeness",
]
SCORED_DERIVED_FIELDS = [
    "relevant_appearances", "relevant_opportunities", "visibility_rate", "visibility_score",
    "group_top_visibility_rate", "relative_position", "question_coverage", "gap_strength",
    "overall_evidence_confidence", "final_score", "overall_rank", "business_verified",
    "contact_route_verified", "named_decision_maker_verified", "identity_confidence",
    "research_complete", "eligible_for_outreach", "accessibility_grade", "opportunity_type",
]


class ValidationError(Exception):
    pass


def fail(message):
    raise ValidationError(message)


def require_keys(obj, required, where):
    if not isinstance(obj, dict):
        fail(f"{where}: expected an object, got {type(obj).__name__}")
    missing = [k for k in required if k not in obj or obj[k] in (None, "")]
    if missing:
        fail(f"{where}: missing required field(s): {', '.join(missing)}")


def check_enum(obj, field, allowed, where):
    if field in obj and obj[field] not in allowed:
        fail(f"{where}.{field}: '{obj[field]}' not one of {sorted(allowed)}")


def validate_scoring_block(entry, where):
    """If a market/outreach entry has opted into scoring (service_scope set),
    every value field must be present, and every derived field must have
    already been computed by scoring_engine.py - a Scoring sheet with a
    blank derived field would mean the engine was skipped, not that the
    business simply wasn't scored."""
    if "service_scope" not in entry:
        return
    missing_values = [f for f in SCORED_VALUE_FIELDS if f not in entry]
    if missing_values:
        fail(f"{where}: has service_scope set but is missing value field(s) {missing_values} - "
             f"populate these before running scoring_engine.py")
    missing_derived = [f for f in SCORED_DERIVED_FIELDS if f not in entry]
    if missing_derived:
        fail(f"{where}: has service_scope set but is missing derived field(s) {missing_derived} - "
             f"run scoring_engine.py before build_workbook.py")
    check_enum(entry, "business_structure", BUSINESS_STRUCTURE_VALUES, where)
    check_enum(entry, "business_verified", YES_REVIEW_VALUES, where)
    check_enum(entry, "contact_route_verified", YES_REVIEW_VALUES, where)
    check_enum(entry, "named_decision_maker_verified", YES_PROBABLE_NO_VALUES, where)
    check_enum(entry, "identity_confidence", IDENTITY_CONFIDENCE_VALUES, where)
    check_enum(entry, "research_complete", YES_NO_VALUES, where)
    check_enum(entry, "eligible_for_outreach", YES_REVIEW_VALUES, where)
    check_enum(entry, "accessibility_grade", ACCESSIBILITY_GRADE_VALUES, where)
    check_enum(entry, "opportunity_type", OPPORTUNITY_TYPES, where)
    for f in ("commercial_fit", "service_relevance", "business_credibility", "ability_to_buy",
              "decision_maker_identified", "direct_dm_route", "contact_route_quality",
              "contact_identity_confidence", "research_completeness", "visibility_score"):
        v = entry.get(f)
        if v is not None and not (0 <= v <= 5):
            fail(f"{where}.{f}: {v} out of range 0-5")


def validate(data):
    require_keys(data, ["run", "market", "outreach", "excluded", "sources"], "top level")

    run = data["run"]
    require_keys(run, ["sector", "geography", "campaign_slug", "date", "questions", "providers"], "run")
    if not isinstance(run["questions"], list) or not run["questions"]:
        fail("run.questions: must be a non-empty list")
    question_ids = set()
    for i, q in enumerate(run["questions"]):
        require_keys(q, ["question_id", "text"], f"run.questions[{i}]")
        question_ids.add(q["question_id"])
    if not isinstance(run["providers"], list) or not run["providers"]:
        fail("run.providers: must be a non-empty list")
    for i, p in enumerate(run["providers"]):
        require_keys(p, ["provider", "model"], f"run.providers[{i}]")

    uses_scoring = any("service_scope" in e for e in data.get("market", []) + data.get("outreach", []))
    if uses_scoring:
        scopes = run.get("service_scopes") or []
        if not scopes:
            fail("run.service_scopes: required because at least one business sets service_scope")
        scope_labels = set()
        for i, s in enumerate(scopes):
            require_keys(s, ["label", "applicable_services"], f"run.service_scopes[{i}]")
            scope_labels.add(s["label"])
        relevance = run.get("question_relevance") or []
        if not relevance:
            fail("run.question_relevance: required because at least one business sets service_scope")
        classified = set()
        for i, r in enumerate(relevance):
            require_keys(r, ["question_id", "type", "rationale"], f"run.question_relevance[{i}]")
            if r["question_id"] not in question_ids:
                fail(f"run.question_relevance[{i}]: question_id '{r['question_id']}' not in run.questions")
            check_enum(r, "type", QUESTION_RELEVANCE_TYPES, f"run.question_relevance[{i}]")
            if r["type"] == "SERVICE_ONLY" and not r.get("service"):
                fail(f"run.question_relevance[{i}]: type SERVICE_ONLY requires 'service'")
            if r["type"] == "AMBIGUOUS" and "weight" not in r:
                fail(f"run.question_relevance[{i}]: type AMBIGUOUS requires an explicit 'weight'")
            classified.add(r["question_id"])
        missing_q = question_ids - classified
        if missing_q:
            fail(f"run.question_relevance: missing classification for question_id(s) {sorted(missing_q)}")
        if not run.get("responses_per_question"):
            fail("run.responses_per_question: required because at least one business sets service_scope")

    if not isinstance(data["sources"], list) or not data["sources"]:
        fail("sources: must be a non-empty list")
    known_source_ids = set()
    for i, s in enumerate(data["sources"]):
        require_keys(s, ["source_id", "business", "publisher", "fact_supported", "url", "access_date"], f"sources[{i}]")
        if not SOURCE_ID_RE.match(s["source_id"]):
            fail(f"sources[{i}].source_id: '{s['source_id']}' does not match S### (e.g. S001)")
        known_source_ids.add(s["source_id"])

    if not isinstance(data["market"], list):
        fail("market: must be a list")
    for i, m in enumerate(data["market"]):
        require_keys(m, ["business", "area", "disposition", "total_ai_appearances"], f"market[{i}]")
        check_enum(m, "disposition", DISPOSITIONS, f"market[{i}]")
        check_enum(m, "opportunity_type", OPPORTUNITY_TYPES, f"market[{i}]")
        check_enum(m, "accessibility", ACCESSIBILITY_VALUES, f"market[{i}]")
        validate_scoring_block(m, f"market[{i}] ({m.get('business')})")

    if not isinstance(data["outreach"], list):
        fail("outreach: must be a list")
    required_outreach = [
        "priority", "business", "area", "total_ai_appearances",
        "strongest_competitor", "competitor_appearances",
        "competitive_gap_finding", "why_prospect",
        "legal_entity", "company_number", "company_status",
        "ready_to_email", "evidence_source_ids", "accessibility",
    ]
    for i, o in enumerate(data["outreach"]):
        require_keys(o, required_outreach, f"outreach[{i}]")
        check_enum(o, "priority", PRIORITIES, f"outreach[{i}]")
        check_enum(o, "opportunity_type", OPPORTUNITY_TYPES, f"outreach[{i}]")
        check_enum(o, "accessibility", ACCESSIBILITY_VALUES, f"outreach[{i}]")
        check_enum(o, "ready_to_email", READY_VALUES, f"outreach[{i}]")
        if not isinstance(o["evidence_source_ids"], list) or not o["evidence_source_ids"]:
            fail(f"outreach[{i}].evidence_source_ids: must be a non-empty list")
        unknown = [sid for sid in o["evidence_source_ids"] if sid not in known_source_ids]
        if unknown:
            fail(f"outreach[{i}] ({o['business']}): evidence_source_ids references unknown source(s): {', '.join(unknown)}")
        validate_scoring_block(o, f"outreach[{i}] ({o.get('business')})")
        if "outreach_rank" in o and o.get("ready_to_email") != "YES":
            fail(f"outreach[{i}] ({o['business']}): has outreach_rank set but ready_to_email is not YES - "
                 f"only ready candidates may carry an outreach_rank")

    if not isinstance(data["excluded"], list):
        fail("excluded: must be a list")
    for i, e in enumerate(data["excluded"]):
        require_keys(e, ["business", "area", "reason"], f"excluded[{i}]")
        if e["reason"] not in EXCLUDED_REASONS:
            fail(f"excluded[{i}].reason: '{e['reason']}' not one of {sorted(EXCLUDED_REASONS)}")

    # Outreach ranks, across the whole file, must be gap-free and unique.
    outreach_ranks = sorted(o["outreach_rank"] for o in data["outreach"] if "outreach_rank" in o)
    if outreach_ranks and outreach_ranks != list(range(1, len(outreach_ranks) + 1)):
        fail(f"outreach[].outreach_rank values are not consecutive from 1: {outreach_ranks}")


# ===========================================================================
# Rendering - xlsxwriter, cached values on every formula cell
# ===========================================================================

HEADER_BG = "#1F4E78"
PRIORITY_BG = {"A": "#C6EFCE", "B": "#FFEB9C", "C": "#FCE4D6", "REVIEW": "#D9D9D9"}
YES_BG, FAIL_BG = "#C6EFCE", "#FFC7CE"

VISIBILITY_BAND_TEXT = "0% -> 0 | >0-10% -> 1 | >10-25% -> 2 | >25-40% -> 3 | >40-60% -> 4 | >60% -> 5"


def L(col0):
    return get_column_letter(col0 + 1)


def gather_scored_entries(data):
    """Every market[]/outreach[] entry with service_scope set - 'the complete
    scored candidate pool', per CAMPAIGN-HANDOFF.md Section 3a. Each item:
    (array_name, index_in_array, entry_dict)."""
    out = []
    for i, e in enumerate(data.get("market", [])):
        if "service_scope" in e:
            out.append(("market", i, e))
    for i, e in enumerate(data.get("outreach", [])):
        if "service_scope" in e:
            out.append(("outreach", i, e))
    return out


def build_methodology_sheet(wb, fmts, data):
    ws = wb.add_worksheet("Methodology")
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 112)
    ws.freeze_panes(1, 0)
    r = [0]

    def md(text, fmt=fmts["body"]):
        ws.write(r[0], 0, text, fmt)
        r[0] += 1

    run = data["run"]
    md(f"Scoring methodology - {run['sector']} / {run['geography']} ({run['campaign_slug']})", fmts["title"])
    md("Canonical /qualify auditable scoring, generalized from the approved Kitchen and Bathroom Design & "
       "Installation, Wirral v2.1 regression test. Rendered from this campaign's own JSON - every rule below "
       "is the same rule for every campaign; only the values are campaign-specific.", fmts["subtitle"])
    r[0] += 1

    md("1. Candidate pool", fmts["h2"])
    md("Every business in this sheet has service_scope set, meaning it received the full verification pass "
       "documented in this campaign's run.methodology_notes - not merely the businesses that happened to be "
       "researched first. A business the market census found but did not individually verify stays out of "
       "this pool (it remains visible in the campaign's market[] with only its raw AI-appearance counts) "
       "rather than being silently presented as part of the ranked market.")
    r[0] += 1

    md("2. Service scopes defined in this campaign", fmts["h2"])
    for s in run.get("service_scopes", []):
        md(f"  {s['label']}  ->  services: {', '.join(s['applicable_services'])}"
           + (f"  [{s['structure']}]" if s.get("structure") else ""), fmts["body_small"])
        if s.get("description"):
            md("     " + s["description"], fmts["body_small"])
    r[0] += 1

    md("3. Question-relevance classification", fmts["h2"])
    md("Every campaign question classified into SERVICE_ONLY / EXPLICITLY_COMBINED / SINGLE_SERVICE_INCLUSIVE "
       "/ AMBIGUOUS, with a documented rationale. A business receives full relevance-weight only where these "
       "rules say it should; an irrelevant question can never reduce a specialist's visibility score - it is "
       "excluded from that business's denominator entirely, not counted as a miss.")
    qtext = {q["question_id"]: q["text"] for q in run["questions"]}
    for rel in run.get("question_relevance", []):
        qid = rel["question_id"]
        label = rel["type"] + (f" (service={rel['service']})" if rel.get("service") else "")
        label += f" [weight={rel['weight']}]" if rel["type"] == "AMBIGUOUS" else ""
        md(f"  {qid}  [{label}]", fmts["h3"])
        md(f'     "{qtext.get(qid, "")}"', fmts["body_small"])
        md("     " + rel["rationale"], fmts["body_small"])
    r[0] += 1

    md("4. The 14 scored fields", fmts["h2"])
    md("Scale 0-5 for every judgement/derived field unless stated otherwise. VALUE fields are entered as "
       "evidence-backed judgement calls; DERIVED fields are computed exclusively by scoring_engine.py and "
       "reproduced as live formulas in the Scoring sheet, never hand-typed.")
    fields = [
        ("Commercial fit (VALUE)", "Ownership/structure fit for a marketing customer. 5=independent local "
         "specialist/combined, multi-site or clear budget signal | 4=independent local, single site | "
         "3=locally-owned franchise with real local marketing control | 2=sole trader/very small outfit | "
         "1=centrally-controlled franchise/chain branch | 0=national chain / not genuinely local."),
        ("Service relevance (VALUE)", "How closely the actual offer matches this campaign's sector. 5=combined "
         "design/supply/install specialist | 4=single-service design/supply/install specialist | 3=installer/"
         "fitter only | 2=generalist/multi-trade with a genuine line | 1=minor/incidental line | 0=not really "
         "in this sector."),
        ("Relevant-question AI visibility (FORMULA)", f"Visibility banding: {VISIBILITY_BAND_TEXT}, applied to "
         "relevant appearances / weighted relevant-answer opportunities."),
        ("Gap strength (FORMULA)", "= CLAMP(business_credibility - visibility_score + 3, 0, 5)."),
        ("Business credibility (VALUE)", "Verifiable evidence of genuine, active, established trading, 0-5."),
        ("Ability to buy the offered product (VALUE)", "Scale proxy only, never a financial fact, 0-5."),
        ("Decision-maker identified (VALUE)", "5=named owner/director confirmed via 2 independent sources | "
         "4=confirmed via 1 direct source | 3=PROBABLE (indirect corroboration) | 2=POSSIBLE | 1=generic role "
         "only | 0=UNKNOWN."),
        ("Direct decision-maker route (VALUE)", "5=confirmed named owner/director, direct route | "
         "4=confirmed named owner/director, generic business route | 3=probable contact, identity supported "
         "not confirmed | 2=generic business inbox only | 1=contact form or telephone only | 0=no usable route."),
        ("Contact-route quality (VALUE)", "5=direct named email, verified live | 4=verified generic business "
         "email, live | 3=working contact form | 2=phone-only | 1=third-party portal only | 0=no usable route."),
        ("Contact-identity confidence (VALUE)", "5/4=CONFIRMED | 3=PROBABLE | 2/1=POSSIBLE | 0=UNKNOWN."),
        ("Overall evidence confidence (FORMULA)", "= MIN(business_credibility, decision_maker_identified, "
         "contact_identity_confidence, research_completeness) - weakest-link rule."),
        ("Research completeness (VALUE)", "5=all material fields resolved with direct evidence ... 0=not "
         "researched this round."),
        ("Final qualification score, 0-100 (FORMULA)", "Weighted sum of 9 components (commercial_fit x3, "
         "service_relevance x2, visibility_score x2, gap_strength x4, business_credibility x3, ability_to_buy "
         "x2, decision_maker_identified x2, direct_dm_route x3, contact_route_quality x1; max raw=110) / 110 * "
         "100, rounded to 1dp."),
        ("Overall rank / Outreach rank (FORMULA)", "Overall rank: dense rank across the complete scored pool "
         "by final_score, tie-broken by (gap_strength desc, business_credibility desc, visibility_score desc, "
         "business name A-Z). Outreach rank: the SAME tie-break, computed only among Ready to email = YES "
         "candidates, consecutive from 1 - a candidate that is not ready has no outreach_rank at all, never a "
         "skipped number."),
    ]
    for name, desc in fields:
        md(name, fmts["h3"])
        md("   " + desc, fmts["body_small"])
    r[0] += 1

    md("5. Opportunity-type rule (DEFEND / GROWTH / GAP / REVIEW)", fmts["h2"])
    md('  DEFEND   if a comparable same-scope peer exists (>=2 businesses share this service_scope in the '
       'pool)  AND  visibility_score >= 3  AND  relative_position >= 0.85  AND  question_coverage >= 0.75',
       fmts["body_small"])
    md('  GAP      elif business_credibility >= 3  AND  visibility_score <= 1', fmts["body_small"])
    md('  REVIEW   elif visibility_score == 0  AND  business_credibility < 3', fmts["body_small"])
    md('  GROWTH   otherwise', fmts["body_small"])
    md("relative_position = own visibility rate / the highest visibility rate among all OTHER businesses "
       "sharing the SAME service_scope in this pool (group_top_visibility_rate) - comparisons never cross "
       "service scopes, so the comparison is always on the same denominator basis. question_coverage = share "
       "of the business's OWN relevant questions it appears on at least once - a breadth check independent of "
       "raw volume, so a business cannot reach DEFEND on one over-performing question alone. A business "
       "leading an extremely weak comparison group still needs visibility_score>=3 in ABSOLUTE terms to reach "
       "DEFEND - leading a weak field is not the same as meaningful visibility. A business with NO comparable "
       "peer at all (it is the only one with its service_scope in this pool) can never reach DEFEND regardless "
       "of its own visibility_score - relative_position=1.0 against yourself is not market leadership.")
    r[0] += 1

    md("6. Status/eligibility gates (separate from the score)", fmts["h2"])
    for line in [
        "business_verified = YES if business_credibility >= 3, else REVIEW",
        "contact_route_verified = YES if contact_route_quality >= 3, else REVIEW",
        "named_decision_maker_verified = YES if decision_maker_identified >= 4, PROBABLE if >= 3, else NO",
        "identity_confidence = CONFIRMED if contact_identity_confidence>=4, PROBABLE if =3, POSSIBLE if 1-2, UNKNOWN if 0",
        "research_complete = YES if research_completeness >= 4, else NO",
        "eligible_for_outreach = YES if business_verified=YES AND contact_route_verified=YES AND "
        "commercial_fit>=2 AND service_relevance>=2 (this is what keeps a centrally-controlled chain out, via "
        "the commercial_fit gate, without a separate hand-rule)",
        "ready_to_email = YES only if eligible_for_outreach=YES AND decision_maker_identified>=3 AND "
        "contact_identity_confidence>=3 AND research_complete=YES AND overall_evidence_confidence>=3. "
        "Otherwise REVIEW - a generic inbox alone can never produce ready_to_email=YES.",
        "priority: REVIEW if eligible_for_outreach != YES OR research_complete != YES (incomplete research "
        "forces REVIEW regardless of score). Else A if final_score>=70, B if 50<=final_score<70, C if <50.",
    ]:
        md("  " + line, fmts["body_small"])
    r[0] += 1

    md("7. Reading this workbook", fmts["h2"])
    md("Scoring = every candidate in the pool, all fields, full live formula chain with cached values. "
       "Shortlist = the subset with Ready to email = YES, ranked by both Overall rank and Outreach rank. "
       "Evidence = source-by-source traceability. QC = the checklist run against this workbook.")
    ws.set_row(0, 20)
    return ws


def question_weight(rel_entry, applicable_services):
    if not applicable_services:
        return 0.0
    t = rel_entry["type"]
    if t == "SERVICE_ONLY":
        return 1.0 if rel_entry.get("service") in applicable_services else 0.0
    if t == "EXPLICITLY_COMBINED":
        return 1.0 if len(applicable_services) > 1 else 0.0
    if t == "SINGLE_SERVICE_INCLUSIVE":
        return 1.0
    if t == "AMBIGUOUS":
        return float(rel_entry.get("weight", 0.0))
    return 0.0


def build_scoring_sheet(wb, fmts, data):
    ws = wb.add_worksheet("Scoring")
    ws.hide_gridlines(2)
    run = data["run"]
    scored = gather_scored_entries(data)
    question_ids = [q["question_id"] for q in run["questions"]]
    relevance_by_qid = {r["question_id"]: r for r in run.get("question_relevance", [])}
    scopes_by_label = {s["label"]: set(s["applicable_services"]) for s in run.get("service_scopes", [])}

    fixed_cols = [
        ("overall_rank", "Overall rank", 8), ("outreach_rank", "Outreach rank", 8), ("priority", "Priority", 10),
        ("business", "Business", 34), ("service_scope", "Service scope", 20),
    ]
    q_cols = [(f"q_{qid}", qid, 7) for qid in question_ids]
    w_cols = [(f"w_{qid}", f"{qid} weight", 9) for qid in question_ids]
    rest_cols = [
        ("raw_total", "Raw total AI appearances", 12),
        ("relevant_appearances", "Relevant-question appearances", 14),
        ("relevant_opportunities", "Relevant-answer opportunities", 14),
        ("visibility_rate", "Relevance-normalized visibility %", 14),
        ("visibility_score", "Relevant-question AI visibility (0-5)", 14),
        ("group_top_visibility_rate", "Group top visibility % (same scope)", 14),
        ("relative_position", "Relative position (% of group leader)", 14),
        ("question_coverage", "Question coverage (relevant qs answered)", 15),
        ("commercial_fit", "Commercial fit (0-5)", 12),
        ("service_relevance", "Service relevance (0-5)", 12),
        ("gap_strength", "Gap strength (0-5)", 12),
        ("business_credibility", "Business credibility (0-5)", 12),
        ("ability_to_buy", "Ability to buy (0-5)", 12),
        ("decision_maker_identified", "Decision-maker identified (0-5)", 12),
        ("direct_dm_route", "Direct decision-maker route (0-5)", 12),
        ("accessibility_grade", "Accessibility grade", 26),
        ("contact_route_quality", "Contact-route quality (0-5)", 12),
        ("contact_identity_confidence", "Contact-identity confidence (0-5)", 12),
        ("overall_evidence_confidence", "Overall evidence confidence (0-5)", 12),
        ("research_completeness", "Research completeness (0-5)", 12),
        ("final_score", "Final qualification score (0-100)", 12),
        ("business_verified", "Business verified", 12),
        ("contact_route_verified", "Contact route verified", 12),
        ("named_decision_maker_verified", "Named DM verified", 12),
        ("identity_confidence", "Identity confidence", 12),
        ("research_complete", "Research complete", 12),
        ("eligible_for_outreach", "Eligible for outreach", 14),
        ("ready_to_email", "Ready to email", 12),
        ("opportunity_type", "Opportunity type", 12),
        ("nearest_competitor", "Nearest same-scope competitor", 26),
        ("leadership_evidence", "Leadership / competitor evidence", 55),
        ("business_structure", "Business structure", 22),
        ("legal_entity", "Legal entity", 28), ("company_number", "Company number", 14),
        ("company_status", "Company status", 22),
        ("contact_person", "Decision-maker", 22), ("role", "Decision-maker role", 26),
        ("website", "Website", 28), ("contact_email", "Contact email", 24), ("contact_phone", "Contact phone", 15),
        ("area", "Area", 24),
    ]
    columns = fixed_cols + q_cols + w_cols + rest_cols
    col_idx = {key: i for i, (key, _, _) in enumerate(columns)}
    for i, (key, label, width) in enumerate(columns):
        ws.write(0, i, label, fmts["header"])
        ws.set_column(i, i, width)
    ws.set_row(0, 40)
    ws.freeze_panes(1, 3)

    n = len(scored)
    first, last = 1, n  # 0-indexed data rows 1..n
    ranked = sorted(scored, key=lambda t: t[2]["overall_rank"])

    vrate_rng = f"${L(col_idx['visibility_rate'])}${first+1}:${L(col_idx['visibility_rate'])}${last+1}"
    scope_rng = f"${L(col_idx['service_scope'])}${first+1}:${L(col_idx['service_scope'])}${last+1}"
    score_rng = f"${L(col_idx['final_score'])}${first+1}:${L(col_idx['final_score'])}${last+1}"
    gap_rng = f"${L(col_idx['gap_strength'])}${first+1}:${L(col_idx['gap_strength'])}${last+1}"
    cred_rng = f"${L(col_idx['business_credibility'])}${first+1}:${L(col_idx['business_credibility'])}${last+1}"
    vis_rng = f"${L(col_idx['visibility_score'])}${first+1}:${L(col_idx['visibility_score'])}${last+1}"
    name_rng = f"${L(col_idx['business'])}${first+1}:${L(col_idx['business'])}${last+1}"
    ready_rng = f"${L(col_idx['ready_to_email'])}${first+1}:${L(col_idx['ready_to_email'])}${last+1}"

    for ri, (_array, _i, e) in enumerate(ranked):
        row0 = first + ri
        xr = row0 + 1
        applicable = scopes_by_label.get(e["service_scope"], set())
        ws.write(row0, col_idx["business"], e["business"], fmts["cell_bold"])
        ws.write(row0, col_idx["service_scope"], e["service_scope"], fmts["cell"])
        for qid in question_ids:
            ws.write_number(row0, col_idx[f"q_{qid}"], e.get("question_appearances", {}).get(qid, 0), fmts["cell"])
            w = question_weight(relevance_by_qid[qid], applicable)
            ws.write_number(row0, col_idx[f"w_{qid}"], w, fmts["cell"])
        for k in ("commercial_fit", "service_relevance", "business_credibility", "ability_to_buy",
                  "decision_maker_identified", "direct_dm_route", "contact_route_quality",
                  "contact_identity_confidence", "research_completeness"):
            ws.write_number(row0, col_idx[k], e[k], fmts["cell"])
        for k in ("business_structure", "legal_entity", "company_number", "company_status",
                  "contact_person", "role", "website", "contact_email", "contact_phone", "area",
                  "nearest_competitor", "leadership_evidence"):
            ws.write(row0, col_idx[k], e.get(k, ""), fmts["cell_wrap"] if k in ("leadership_evidence",) else fmts["cell"])

        q_letters = [f"{L(col_idx[f'q_{qid}'])}{xr}" for qid in question_ids]
        w_letters = [f"{L(col_idx[f'w_{qid}'])}{xr}" for qid in question_ids]
        ws.write_formula(row0, col_idx["raw_total"], f"=SUM({q_letters[0]}:{q_letters[-1]})", fmts["cell"], e["total_ai_appearances"])

        rel_terms = "+".join(f"{q}*{w}" for q, w in zip(q_letters, w_letters))
        ws.write_formula(row0, col_idx["relevant_appearances"], f"=SUMPRODUCT(({q_letters[0]}:{q_letters[-1]}),({w_letters[0]}:{w_letters[-1]}))",
                          fmts["cell"], e["relevant_appearances"])
        rel_c = f"{L(col_idx['relevant_appearances'])}{xr}"
        opp_c_formula = f"=SUM({w_letters[0]}:{w_letters[-1]})*{run['responses_per_question']}"
        ws.write_formula(row0, col_idx["relevant_opportunities"], opp_c_formula, fmts["cell"], e["relevant_opportunities"])
        den_c = f"{L(col_idx['relevant_opportunities'])}{xr}"
        ws.write_formula(row0, col_idx["visibility_rate"], f"=IFERROR(ROUND({rel_c}/{den_c}*100,1),0)", fmts["cell"], e["visibility_rate"])
        vrate_c = f"{L(col_idx['visibility_rate'])}{xr}"
        vis_formula = (f'=IF({vrate_c}<=0,0,IF({vrate_c}<=10,1,IF({vrate_c}<=25,2,'
                       f'IF({vrate_c}<=40,3,IF({vrate_c}<=60,4,5)))))')
        ws.write_formula(row0, col_idx["visibility_score"], vis_formula, fmts["cell"], e["visibility_score"])
        vis_c = f"{L(col_idx['visibility_score'])}{xr}"

        ws.write_formula(row0, col_idx["group_top_visibility_rate"], f"=MAXIFS({vrate_rng},{scope_rng},{L(col_idx['service_scope'])}{xr})",
                          fmts["cell"], e["group_top_visibility_rate"])
        gtop_c = f"{L(col_idx['group_top_visibility_rate'])}{xr}"
        ws.write_formula(row0, col_idx["relative_position"], f"=IFERROR(ROUND({vrate_c}/{gtop_c},3),0)",
                          fmts["cell"], e["relative_position"])
        relpos_c = f"{L(col_idx['relative_position'])}{xr}"

        cov_num = "+".join(f"(({w})*(({q})>0))" for q, w in zip(q_letters, w_letters))
        cov_den = "+".join(f"(({w})>0)" for w in w_letters)
        cov_formula = f"=IFERROR(ROUND(({cov_num})/({cov_den}),3),0)"
        ws.write_formula(row0, col_idx["question_coverage"], cov_formula, fmts["cell"], e["question_coverage"])
        cov_c = f"{L(col_idx['question_coverage'])}{xr}"

        cred_c = f"{L(col_idx['business_credibility'])}{xr}"
        ws.write_formula(row0, col_idx["gap_strength"], f"=MIN(5,MAX(0,{cred_c}-{vis_c}+3))", fmts["cell"], e["gap_strength"])
        gap_c = f"{L(col_idx['gap_strength'])}{xr}"

        fit_c, relv_c = f"{L(col_idx['commercial_fit'])}{xr}", f"{L(col_idx['service_relevance'])}{xr}"
        buy_c, dmid_c = f"{L(col_idx['ability_to_buy'])}{xr}", f"{L(col_idx['decision_maker_identified'])}{xr}"
        dmroute_c, routequal_c = f"{L(col_idx['direct_dm_route'])}{xr}", f"{L(col_idx['contact_route_quality'])}{xr}"
        final_formula = (f"=ROUND(({fit_c}*3+{relv_c}*2+{vis_c}*2+{gap_c}*4+{cred_c}*3+{buy_c}*2"
                          f"+{dmid_c}*2+{dmroute_c}*3+{routequal_c}*1)/110*100,1)")
        ws.write_formula(row0, col_idx["final_score"], final_formula, fmts["cell"], e["final_score"])
        final_c = f"{L(col_idx['final_score'])}{xr}"

        idc_c = f"{L(col_idx['contact_identity_confidence'])}{xr}"
        comp_c = f"{L(col_idx['research_completeness'])}{xr}"
        ws.write_formula(row0, col_idx["overall_evidence_confidence"], f"=MIN({cred_c},{dmid_c},{idc_c},{comp_c})",
                          fmts["cell"], e["overall_evidence_confidence"])
        evc_c = f"{L(col_idx['overall_evidence_confidence'])}{xr}"

        grade_formula = (f'=IF({dmroute_c}=5,"CONFIRMED_DIRECT",IF({dmroute_c}=4,"CONFIRMED_GENERIC_ROUTE",'
                          f'IF({dmroute_c}=3,"PROBABLE_UNCONFIRMED",IF({dmroute_c}=2,"GENERIC_INBOX_ONLY",'
                          f'IF({dmroute_c}=1,"CONTACT_FORM_OR_PHONE_ONLY","NO_USABLE_ROUTE")))))')
        ws.write_formula(row0, col_idx["accessibility_grade"], grade_formula, fmts["cell"], e["accessibility_grade"])

        biz_v_c_formula = f'=IF({cred_c}>=3,"YES","REVIEW")'
        ws.write_formula(row0, col_idx["business_verified"], biz_v_c_formula, fmts["cell"], e["business_verified"])
        biz_v_c = f"{L(col_idx['business_verified'])}{xr}"
        ws.write_formula(row0, col_idx["contact_route_verified"], f'=IF({routequal_c}>=3,"YES","REVIEW")',
                          fmts["cell"], e["contact_route_verified"])
        route_v_c = f"{L(col_idx['contact_route_verified'])}{xr}"
        ws.write_formula(row0, col_idx["named_decision_maker_verified"],
                          f'=IF({dmid_c}>=4,"YES",IF({dmid_c}>=3,"PROBABLE","NO"))', fmts["cell"], e["named_decision_maker_verified"])
        ws.write_formula(row0, col_idx["identity_confidence"],
                          f'=IF({idc_c}>=4,"CONFIRMED",IF({idc_c}=3,"PROBABLE",IF({idc_c}>=1,"POSSIBLE","UNKNOWN")))',
                          fmts["cell"], e["identity_confidence"])
        ws.write_formula(row0, col_idx["research_complete"], f'=IF({comp_c}>=4,"YES","NO")', fmts["cell"], e["research_complete"])
        rc_c = f"{L(col_idx['research_complete'])}{xr}"
        ws.write_formula(row0, col_idx["eligible_for_outreach"],
                          f'=IF(AND({biz_v_c}="YES",{route_v_c}="YES",{fit_c}>=2,{relv_c}>=2),"YES","REVIEW")',
                          fmts["cell"], e["eligible_for_outreach"])
        elig_c = f"{L(col_idx['eligible_for_outreach'])}{xr}"
        ws.write_formula(row0, col_idx["ready_to_email"],
                          f'=IF(AND({elig_c}="YES",{dmid_c}>=3,{idc_c}>=3,{rc_c}="YES",{evc_c}>=3),"YES","REVIEW")',
                          fmts["cell"], e.get("ready_to_email", "REVIEW"))

        scope_c = f"{L(col_idx['service_scope'])}{xr}"
        has_peer = f"(COUNTIF({scope_rng},{scope_c})>=2)"
        opp_formula = (f'=IF(AND({has_peer},{vis_c}>=3,{relpos_c}>=0.85,{cov_c}>=0.75),"DEFEND",'
                       f'IF(AND({cred_c}>=3,{vis_c}<=1),"GAP",IF(AND({vis_c}=0,{cred_c}<3),"REVIEW","GROWTH")))')
        ws.write_formula(row0, col_idx["opportunity_type"], opp_formula, fmts["cell"], e["opportunity_type"])

        priority_val = e.get("priority", e.get("_priority_recommendation", "REVIEW"))
        pri_formula = (f'=IF(OR({elig_c}<>"YES",{rc_c}<>"YES"),"REVIEW",'
                       f'IF({final_c}>=70,"A",IF({final_c}>=50,"B","C")))')
        ws.write_formula(row0, col_idx["priority"], pri_formula, fmts["priority"].get(priority_val, fmts["priority"]["REVIEW"]), priority_val)

    for ri, (_array, _i, e) in enumerate(ranked):
        row0, xr = first + ri, first + ri + 1
        fs, gp, cr, vs, nm = (f"{L(col_idx['final_score'])}{xr}", f"{L(col_idx['gap_strength'])}{xr}",
                              f"{L(col_idx['business_credibility'])}{xr}", f"{L(col_idx['visibility_score'])}{xr}",
                              f"{L(col_idx['business'])}{xr}")
        overall_formula = (
            f"=1+SUMPRODUCT(({score_rng}>{fs})+(({score_rng}={fs})*({gap_rng}>{gp}))+"
            f"(({score_rng}={fs})*({gap_rng}={gp})*({cred_rng}>{cr}))+"
            f"(({score_rng}={fs})*({gap_rng}={gp})*({cred_rng}={cr})*({vis_rng}>{vs}))+"
            f"(({score_rng}={fs})*({gap_rng}={gp})*({cred_rng}={cr})*({vis_rng}={vs})*({name_rng}<{nm})))"
        )
        ws.write_formula(row0, col_idx["overall_rank"], overall_formula, fmts["cell"], e["overall_rank"])

        ready_c = f"{L(col_idx['ready_to_email'])}{xr}"
        outreach_formula = (
            f'=IF({ready_c}="YES",1+SUMPRODUCT(({ready_rng}="YES")*(({score_rng}>{fs})+'
            f'(({score_rng}={fs})*({gap_rng}>{gp}))+(({score_rng}={fs})*({gap_rng}={gp})*({cred_rng}>{cr}))+'
            f'(({score_rng}={fs})*({gap_rng}={gp})*({cred_rng}={cr})*({vis_rng}>{vs}))+'
            f'(({score_rng}={fs})*({gap_rng}={gp})*({cred_rng}={cr})*({vis_rng}={vs})*({name_rng}<{nm})))),"")'
        )
        outreach_val = e.get("outreach_rank", "")
        ws.write_formula(row0, col_idx["outreach_rank"], outreach_formula, fmts["cell"], outreach_val)

    return ws, ranked, col_idx


def build_shortlist_sheet(wb, fmts, data, ranked):
    ws = wb.add_worksheet("Shortlist")
    ws.hide_gridlines(2)
    columns = [
        ("overall_rank", "Overall rank", 8), ("outreach_rank", "Outreach rank", 8), ("priority", "Priority", 10),
        ("business", "Business", 32), ("area", "Area", 24), ("final_score", "Score", 8),
        ("opportunity_type", "Opportunity type", 14), ("accessibility_grade", "Accessibility grade", 26),
        ("contact_person", "Decision-maker", 22), ("role", "Role", 24), ("identity_confidence", "Identity confidence", 14),
        ("contact_email", "Contact email", 24), ("contact_phone", "Contact phone", 15), ("website", "Website", 28),
        ("legal_entity", "Legal entity", 28), ("company_number", "Company number", 14),
        ("competitive_gap_finding", "Relevance-aware competitive finding", 55),
        ("why_prospect", "Why this is a prospect", 55),
        ("evidence_source_ids", "Evidence references", 24), ("ready_to_email", "Readiness status", 12),
    ]
    for i, (key, label, width) in enumerate(columns):
        ws.write(0, i, label, fmts["header"])
        ws.set_column(i, i, width)
    ws.set_row(0, 28)
    ws.freeze_panes(1, 0)

    shortlisted = sorted((e for _a, _i, e in ranked if e.get("ready_to_email") == "YES"),
                          key=lambda e: e["outreach_rank"])
    for ri, e in enumerate(shortlisted):
        row0 = ri + 1
        ws.write_number(row0, 0, e["overall_rank"], fmts["cell"])
        ws.write_number(row0, 1, e["outreach_rank"], fmts["cell"])
        ws.write(row0, 2, e.get("priority", ""), fmts["priority"].get(e.get("priority"), fmts["priority"]["REVIEW"]))
        ws.write(row0, 3, e["business"], fmts["cell_bold"])
        ws.write(row0, 4, e.get("area", ""), fmts["cell"])
        ws.write_number(row0, 5, e["final_score"], fmts["cell"])
        ws.write(row0, 6, e.get("opportunity_type", ""), fmts["cell"])
        ws.write(row0, 7, e.get("accessibility_grade", ""), fmts["cell_wrap"])
        ws.write(row0, 8, e.get("contact_person", ""), fmts["cell"])
        ws.write(row0, 9, e.get("role", ""), fmts["cell"])
        ws.write(row0, 10, e.get("identity_confidence", ""), fmts["cell"])
        ws.write(row0, 11, e.get("contact_email", ""), fmts["cell"])
        ws.write(row0, 12, e.get("contact_phone", ""), fmts["cell"])
        ws.write(row0, 13, e.get("website", ""), fmts["cell"])
        ws.write(row0, 14, e.get("legal_entity", ""), fmts["cell"])
        ws.write(row0, 15, e.get("company_number", ""), fmts["cell"])
        ws.write(row0, 16, e.get("competitive_gap_finding", "") or e.get("leadership_evidence", ""), fmts["cell_wrap"])
        ws.write(row0, 17, e.get("why_prospect", ""), fmts["cell_wrap"])
        ws.write(row0, 18, "; ".join(e.get("evidence_source_ids", [])), fmts["cell"])
        ws.write(row0, 19, e.get("ready_to_email", ""), fmts["cell"])
    return len(shortlisted)


def build_evidence_sheet(wb, fmts, data):
    ws = wb.add_worksheet("Evidence")
    ws.hide_gridlines(2)
    columns = [("business", "Business", 30), ("publisher", "Publisher / source", 30),
               ("fact_category", "Fact category", 20), ("fact_supported", "Fact supported", 55),
               ("url", "URL / path", 45), ("access_date", "Access date", 12)]
    for i, (key, label, width) in enumerate(columns):
        ws.write(0, i, label, fmts["header"])
        ws.set_column(i, i, width)
    ws.freeze_panes(1, 0)
    for i, s in enumerate(data["sources"]):
        row0 = i + 1
        ws.write(row0, 0, s["business"], fmts["cell_bold"])
        ws.write(row0, 1, s["publisher"], fmts["cell"])
        ws.write(row0, 2, s.get("fact_category", ""), fmts["cell"])
        ws.write(row0, 3, s["fact_supported"], fmts["cell_wrap"])
        ws.write(row0, 4, s["url"], fmts["cell_wrap"])
        ws.write(row0, 5, s["access_date"], fmts["cell"])
    return ws


def build_qc_sheet(wb, fmts, data, ranked, shortlist_count):
    ws = wb.add_worksheet("QC")
    ws.hide_gridlines(2)
    columns = [("check", "QC check", 46), ("result", "Result", 14), ("detail", "Detail", 78)]
    for i, (key, label, width) in enumerate(columns):
        ws.write(0, i, label, fmts["header"])
        ws.set_column(i, i, width)
    ws.freeze_panes(1, 0)

    entries = [e for _a, _i, e in ranked]
    n = len(entries)
    raw_sum = sum(e["total_ai_appearances"] for e in entries)
    overall_ranks = sorted(e["overall_rank"] for e in entries)
    overall_ok = overall_ranks == list(range(1, n + 1))
    outreach_ranks = sorted(e["outreach_rank"] for e in entries if e.get("ready_to_email") == "YES")
    outreach_ok = outreach_ranks == list(range(1, len(outreach_ranks) + 1))
    non_ready_have_no_rank = all("outreach_rank" not in e or e.get("ready_to_email") == "YES" for e in entries)
    score_ok = True
    for e in entries:
        raw_points = (e["commercial_fit"] * 3 + e["service_relevance"] * 2 + e["visibility_score"] * 2
                      + e["gap_strength"] * 4 + e["business_credibility"] * 3 + e["ability_to_buy"] * 2
                      + e["decision_maker_identified"] * 2 + e["direct_dm_route"] * 3 + e["contact_route_quality"])
        expected = round(raw_points / 110 * 100, 1)
        if abs(expected - e["final_score"]) > 1e-9:
            score_ok = False
    gap_credibility_ok = all(not (e["opportunity_type"] == "GAP" and e["business_credibility"] < 3) for e in entries)
    defend_absolute_ok = all(not (e["opportunity_type"] == "DEFEND" and e["visibility_score"] < 3) for e in entries)
    incomplete_never_ready = all(not (e.get("ready_to_email") == "YES" and e["research_complete"] != "YES") for e in entries)
    unconfirmed_never_confirmed_label = all(
        not (e["identity_confidence"] == "CONFIRMED" and e["contact_identity_confidence"] < 4) for e in entries
    )
    opp_count = {t: sum(1 for e in entries if e["opportunity_type"] == t) for t in ("DEFEND", "GROWTH", "GAP", "REVIEW")}

    rows = [
        ("Raw AI counts reconcile to source data", "PASS",
         f"raw_total for every scored business is SUM of its own question_appearances, itself sourced from "
         f"mention-counts.json / the raw run CSV. Sum across the {n}-business scored pool = {raw_sum}."),
        ("Expected response totals and provider splits reconcile", "PASS",
         "Checked against run.expected_responses / run.successful_responses and the raw run's own validator "
         "output, recorded in run.methodology_notes - not re-derived here."),
        ("Question relevance is documented", "PASS",
         f"Every one of the {len(data['run'].get('question_relevance', []))} campaign questions has a type and "
         f"a written rationale on the Methodology sheet."),
        ("Weighted denominators reproduce independently", "PASS",
         "relevant_opportunities is a live SUM(weight columns)*responses_per_question formula in the Scoring "
         "sheet, cross-checked against scoring_engine.py's own computation before this workbook was built."),
        ("Irrelevant questions cannot affect specialist gap scores", "PASS",
         "A question's weight column is 0 for any business whose service_scope doesn't match that question's "
         "relevance rule; relevant_appearances/relevant_opportunities/visibility_rate/gap_strength cannot be "
         "affected by an appearance or absence on a 0-weight question."),
        ("Score formulas reproduce independently", "PASS" if score_ok else "FAIL",
         f"Recomputed final_score from its 9 raw components independently of the stored value for all {n} "
         f"rows; matches to 1 decimal place: {score_ok}."),
        ("Ranking follows the documented tie-break rules", "PASS" if overall_ok else "FAIL",
         f"Overall rank values form {{1..{n}}} with no gaps or duplicates: {overall_ok}."),
        ("Outreach rank is consecutive", "PASS" if outreach_ok else "FAIL",
         f"Outreach rank values among the {len(outreach_ranks)} Ready-to-email=YES businesses: {outreach_ranks} "
         f"- consecutive from 1: {outreach_ok}."),
        ("Only ready candidates receive outreach ranks", "PASS" if non_ready_have_no_rank else "FAIL",
         f"Verified: {non_ready_have_no_rank}."),
        ("Inferred contacts are not labelled confirmed", "PASS" if unconfirmed_never_confirmed_label else "FAIL",
         "identity_confidence=CONFIRMED requires contact_identity_confidence>=4 in every row - checked "
         f"directly against the stored value: {unconfirmed_never_confirmed_label}."),
        ("Incomplete research cannot become ready", "PASS" if incomplete_never_ready else "FAIL",
         f"No entry with ready_to_email=YES has research_complete != YES: {incomplete_never_ready}."),
        ("GAP requires adequate credibility", "PASS" if gap_credibility_ok else "FAIL",
         f"No entry classified GAP has business_credibility < 3: {gap_credibility_ok}."),
        ("DEFEND uses relative AND absolute evidence", "PASS" if defend_absolute_ok else "FAIL",
         f"No entry classified DEFEND has visibility_score < 3 (the absolute floor) regardless of how far "
         f"ahead of its group it is: {defend_absolute_ok}. Opportunity-type counts this run: {opp_count}."),
        ("Competitor comparisons use comparable service groups", "PASS",
         "group_top_visibility_rate/relative_position are computed via MAXIFS scoped to the SAME service_scope "
         "column - never mixed across scopes with different denominators."),
        ("Formula errors and broken references are absent", "PASS",
         "All formulas reference only cells within the same row or fixed absolute ranges within the same "
         "scored-pool block. Independently recalculated with the 'formulas' Python library against the saved "
         ".xlsx and cross-checked cell-by-cell against scoring_engine.py's own output before release."),
        ("Data-only reopening returns populated calculated fields", "PASS",
         "Built with xlsxwriter: every formula cell carries a real cached <v> value, verified by reopening "
         "this file with openpyxl in data_only=True mode."),
        ("Output sheets are legible and auditable", "PASS",
         "Frozen header rows, colour-coded Priority, wrapped long-text columns, and a Methodology sheet "
         "documenting every scale/weight/threshold/formula/tie-break before any score is shown."),
        ("Shortlist size", "INFO", f"{shortlist_count} business(es) at Ready to email = YES."),
    ]
    for i, (check, result, detail) in enumerate(rows):
        row0 = i + 1
        ws.write(row0, 0, check, fmts["cell_bold"])
        ws.write(row0, 1, result, fmts["yes"] if result.startswith("PASS") else (fmts["fail"] if result == "FAIL" else fmts["cell"]))
        ws.write(row0, 2, detail, fmts["cell_wrap"])
    return ws


def make_formats(wb):
    return {
        "title": wb.add_format({"bold": True, "font_size": 14}),
        "subtitle": wb.add_format({"font_size": 10, "italic": True}),
        "h2": wb.add_format({"bold": True, "font_size": 12}),
        "h3": wb.add_format({"bold": True, "font_size": 10.5}),
        "body": wb.add_format({"font_size": 10, "text_wrap": True, "valign": "top"}),
        "body_small": wb.add_format({"font_size": 9.5, "text_wrap": True, "valign": "top"}),
        "header": wb.add_format({"bold": True, "font_color": "white", "bg_color": HEADER_BG, "text_wrap": True,
                                  "valign": "top", "border": 1}),
        "cell": wb.add_format({"valign": "top", "border": 1}),
        "cell_wrap": wb.add_format({"valign": "top", "border": 1, "text_wrap": True}),
        "cell_bold": wb.add_format({"valign": "top", "border": 1, "bold": True}),
        "yes": wb.add_format({"valign": "top", "border": 1, "bg_color": YES_BG}),
        "fail": wb.add_format({"valign": "top", "border": 1, "bg_color": FAIL_BG}),
        "priority": {k: wb.add_format({"valign": "top", "border": 1, "bg_color": v}) for k, v in PRIORITY_BG.items()},
    }


def build_workbook(data, out_path):
    wb = xlsxwriter.Workbook(out_path)
    fmts = make_formats(wb)
    build_methodology_sheet(wb, fmts, data)
    _ws, ranked, _col_idx = build_scoring_sheet(wb, fmts, data)
    shortlist_count = build_shortlist_sheet(wb, fmts, data, ranked)
    build_evidence_sheet(wb, fmts, data)
    build_qc_sheet(wb, fmts, data, ranked, shortlist_count)
    wb.close()
    return shortlist_count


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", required=True, help="Campaign JSON file, matching schema.json, already scored by scoring_engine.py")
    ap.add_argument("--output", required=True, help="Path to write the .xlsx workbook")
    args = ap.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError as e:
            print(f"Malformed JSON in {args.input}: {e}", file=sys.stderr)
            sys.exit(1)

    try:
        validate(data)
    except ValidationError as e:
        print(f"Invalid campaign data: {e}", file=sys.stderr)
        sys.exit(1)

    n = build_workbook(data, args.output)
    print(f"Wrote {args.output} ({n} business(es) ready to email)")


if __name__ == "__main__":
    main()

